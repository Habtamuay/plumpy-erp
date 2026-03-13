from django.contrib.auth.decorators import login_required
from django.shortcuts import redirect, render
from django.utils import timezone
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

from apps.company.models import Company

from .models import FiscalPeriod
from .services.report_service import (
    get_balance_sheet,
    get_cash_flow,
    get_profit_and_loss,
    get_trial_balance,
)


def _current_company(request):
    company = getattr(request, 'company', None)
    if company:
        return company

    company_id = request.session.get('current_company_id')
    if not company_id:
        return None
    return Company.objects.filter(id=company_id, is_active=True).first()


@login_required
def financial_trial_balance(request):
    company = _current_company(request)
    if not company:
        return redirect('core:home')

    end_date_str = request.GET.get('date')
    end_date = None
    if end_date_str:
        try:
            end_date = timezone.datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            pass

    report = get_trial_balance(company) # Trial balance usually is current, or we could add date support too. keeping simple as requested.
    
    return render(
        request,
        'accounting/financial_trial_balance.html',
        {'report': report, 'company': company, 'today': timezone.now().date()},
    )


@login_required
def financial_profit_and_loss(request):
    company = _current_company(request)
    if not company:
        return redirect('core:home')

    # Handle filters
    period_id = request.GET.get('period')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    start_date, end_date = None, None
    selected_period = None
    today = timezone.now().date()

    fiscal_periods = FiscalPeriod.objects.filter(company=company).order_by('-start_date')

    if period_id and period_id.isdigit():
        try:
            period = FiscalPeriod.objects.get(id=period_id, company=company)
            start_date = period.start_date
            end_date = period.end_date
            selected_period = int(period.id)
        except FiscalPeriod.DoesNotExist:
            pass  # or messages.error
    elif start_date_str and end_date_str:
        try:
            start_date = timezone.datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = timezone.datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            pass

    # if no filters, default to current fiscal period to date
    if not (start_date and end_date):
        current_period = fiscal_periods.filter(start_date__lte=today, end_date__gte=today, is_open=True).first()
        if current_period:
            start_date = current_period.start_date
            end_date = today
            selected_period = current_period.id
        else:
            # Fallback to current month to date if no active period found
            start_date = today.replace(day=1)
            end_date = today

    # Calculate YoY Comparison Dates
    compare_start_date = None
    compare_end_date = None
    if start_date and end_date:
        try:
            compare_start_date = start_date.replace(year=start_date.year - 1)
        except ValueError: # Handle leap year Feb 29
            compare_start_date = start_date.replace(month=2, day=28, year=start_date.year - 1)
        
        try:
            compare_end_date = end_date.replace(year=end_date.year - 1)
        except ValueError:
            compare_end_date = end_date.replace(month=2, day=28, year=end_date.year - 1)

    report = get_profit_and_loss(company, start_date=start_date, end_date=end_date, 
                                 compare_start_date=compare_start_date, compare_end_date=compare_end_date)

    if request.GET.get('export') == 'excel':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"Profit_Loss_{company.name}_{end_date}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Profit and Loss"

        # Styles
        bold_font = Font(bold=True)
        header_font = Font(bold=True, size=14)
        
        # Header
        ws.append([company.name])
        ws['A1'].font = header_font
        ws.append(["Profit and Loss Statement"])
        ws.append([f"Period: {start_date} to {end_date}"])
        ws.append([f"Comparison: {compare_start_date} to {compare_end_date}"])
        ws.append([])
        
        # Header Row
        ws.append(["Account", "Current Period", "Budget", "Variance", "Previous Year"])
        for cell in ws[ws.max_row]:
            cell.font = bold_font

        def write_section(title, data):
            ws.append([title])
            ws[f'A{ws.max_row}'].font = bold_font
            
            for group in data:
                ws.append([group['name']])
                ws[f'A{ws.max_row}'].font = Font(bold=True, italic=True)
                for acc in group['accounts']:
                    ws.append([
                        f"    {acc['name']}", 
                        acc['balance'], 
                        acc['budget'],
                        acc['variance'],
                        acc['previous_balance']
                    ])
                ws.append([f"  Total {group['name']}", group['total'], group['budget_total'], group['variance_total'], group['previous_total']])
                ws[f'A{ws.max_row}'].font = Font(italic=True)
            ws.append([])

        write_section("REVENUES", report['revenues'])
        ws.append(["TOTAL REVENUE", report['total_revenue'], report['total_budget_revenue'], report['total_revenue_variance'], report['total_prev_revenue']])
        ws[f'A{ws.max_row}'].font = bold_font
        ws.append([])

        write_section("EXPENSES", report['expenses'])
        ws.append(["TOTAL EXPENSES", report['total_expenses'], report['total_budget_expenses'], report['total_expenses_variance'], report['total_prev_expenses']])
        ws[f'A{ws.max_row}'].font = bold_font
        ws.append([])

        ws.append(["NET INCOME", report['net_income'], report['budget_net_income'], report['net_income_variance'], report['prev_net_income']])
        ws[f'A{ws.max_row}'].font = header_font

        wb.save(response)
        return response

    context = {
        'report': report, 
        'company': company, 
        'start_date': start_date, 
        'end_date': end_date, 
        'compare_start_date': compare_start_date,
        'compare_end_date': compare_end_date,
        'fiscal_periods': fiscal_periods, 
        'selected_period': selected_period
    }
    return render(
        request,
        'accounting/financial_profit_and_loss.html',
        context,
    )


@login_required
def financial_balance_sheet(request):
    company = _current_company(request)
    if not company:
        return redirect('core:home')

    end_date_str = request.GET.get('date')
    end_date = None
    if end_date_str:
        try:
            end_date = timezone.datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            pass

    report = get_balance_sheet(company, end_date=end_date)
    return render(
        request,
        'accounting/financial_balance_sheet.html',
        {'report': report, 'company': company, 'end_date': end_date},
    )


@login_required
def financial_cash_flow(request):
    company = _current_company(request)
    if not company:
        return redirect('core:home')

    report = get_cash_flow(company)
    return render(
        request,
        'accounting/financial_cash_flow.html',
        {'report': report, 'company': company},
    )
