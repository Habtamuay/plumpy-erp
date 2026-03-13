from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q, Sum, Count, F
from django.db import transaction
from django.utils import timezone
from django.contrib import messages
from django.http import JsonResponse, FileResponse
from django.core.exceptions import ValidationError
from django.urls import reverse
from django.views.decorators.http import require_POST
import json
from decimal import Decimal, InvalidOperation
import csv
from datetime import timedelta

from .utils import generate_po_pdf
from apps.company.models import Company

from .models import (
    PurchaseOrder, PurchaseOrderLine, PurchaseOrderApproval, Supplier,
    PurchaseRequisition, PurchaseRequisitionLine, GoodsReceipt, GoodsReceiptLine,
    VendorPerformance
)
from apps.inventory.models import Item, Warehouse, Lot, StockTransaction
from apps.core.models import Unit


# ============================
# Local Helpers
# ============================

def _generate_auto_item_code():
    """Generate a unique auto item code for PO quick-create."""
    base = timezone.now().strftime("AUTO-%Y%m%d")
    seq = 1
    code = f"{base}-{seq:03d}"
    while Item.objects.filter(code=code).exists():
        seq += 1
        code = f"{base}-{seq:03d}"
    return code


def _resolve_po_item(item_token, unit_id=None):
    """
    Resolve item token posted from PO form.
    Supports:
    - existing item id, e.g. "12"
    - new tag value, e.g. "new:Item Name"
    """
    token = (item_token or '').strip()
    if not token:
        raise ValueError("Empty item token")

    if token.isdigit():
        return Item.objects.get(id=int(token))

    if token.startswith('new:'):
        item_name = token[4:].strip()
        if not item_name:
            raise ValueError("Invalid item name")

        unit = None
        if unit_id:
            try:
                unit = Unit.objects.get(id=int(unit_id))
            except (Unit.DoesNotExist, ValueError, TypeError):
                unit = None
        if not unit:
            unit = Unit.objects.order_by('id').first()
        if not unit:
            raise ValueError("No Unit configured in system")

        return Item.objects.create(
            code=_generate_auto_item_code(),
            name=item_name,
            category='raw',
            unit=unit,
            is_active=True,
            is_purchased=True,
            is_sold=False,
            unit_cost=Decimal('0.00'),
            current_stock=Decimal('0.00'),
            minimum_stock=Decimal('0.00'),
            reorder_point=Decimal('0.00'),
        )

    raise ValueError("Unsupported item token")


def _calculate_withholding_tax(withholding_base, currency):
    """
    Ethiopian WHT rule for this system:
    - 3% only when currency is ETB
    - and taxable base is above 20,000 ETB
    """
    base = withholding_base or Decimal('0.00')
    if currency == 'ETB' and base > Decimal('20000'):
        return (base * Decimal('0.03')).quantize(Decimal('0.01'))
    return Decimal('0.00')


# ============================
# Dashboard Views
# ============================

@login_required
def dashboard(request):
    """
    Purchasing dashboard view
    """
    company_name = request.session.get('current_company_name')
    
    # Base queryset with company filter
    po_qs = PurchaseOrder.objects.all()
    if company_name:
        po_qs = po_qs.filter(company=company_name)
    
    # Get counts
    total_po_count = po_qs.count()
    pending_po_count = po_qs.filter(status='pending').count()
    approved_po_count = po_qs.filter(status='approved').count()
    ordered_po_count = po_qs.filter(status='ordered').count()
    received_po_count = po_qs.filter(status='received').count()
    
    # Get recent POs
    recent_pos = po_qs.select_related('supplier').order_by('-order_date')[:5]
    
    # Get pending approvals
    pending_approvals = PurchaseOrderApproval.objects.filter(
        status='pending',
        po__in=po_qs
    ).select_related('po', 'approver')[:5]
    
    # Get recent requisitions
    req_qs = PurchaseRequisition.objects.all()
    if company_name:
        req_qs = req_qs.filter(company=company_name)
    recent_requisitions = req_qs.order_by('-requested_date')[:5]
    
    # Get total spend this month
    this_month_start = timezone.now().replace(day=1)
    monthly_spend = po_qs.filter(
        status='received',
        received_date__gte=this_month_start
    ).aggregate(total=Sum('total_amount'))['total'] or 0
    
    context = {
        'current_period': timezone.now().strftime('%B %Y'),
        'total_po_count': total_po_count,
        'pending_po_count': pending_po_count,
        'approved_po_count': approved_po_count,
        'ordered_po_count': ordered_po_count,
        'received_po_count': received_po_count,
        'total_suppliers': Supplier.objects.filter(is_active=True).count(),
        'recent_pos': recent_pos,
        'pending_approvals': pending_approvals,
        'recent_requisitions': recent_requisitions,
        'monthly_spend': monthly_spend,
    }
    return render(request, 'purchasing/dashboard.html', context)


@login_required
def supplier_dashboard(request):
    """
    Supplier dashboard view
    """
    company_name = request.session.get('current_company_name')
    suppliers = Supplier.objects.filter(is_active=True)
    if company_name:
        suppliers = suppliers.filter(company=company_name)
    
    context = {
        'suppliers': suppliers,
        'total_suppliers': suppliers.count(),
    }
    return render(request, 'purchasing/supplier_dashboard.html', context)


# ============================
# Purchase Order Views
# ============================

@login_required
def po_list(request):
    """
    List all purchase orders with filtering
    """
    # Base queryset
    queryset = PurchaseOrder.objects.select_related(
        'supplier', 'created_by'
    ).prefetch_related('lines').all()

    # Scope by selected company from session
    company_name = request.session.get('current_company_name')
    if company_name:
        queryset = queryset.filter(company=company_name)

    # Apply filters
    status = request.GET.get('status')
    if status:
        queryset = queryset.filter(status=status)

    supplier_id = request.GET.get('supplier')
    if supplier_id:
        queryset = queryset.filter(supplier_id=supplier_id)

    search = request.GET.get('search')
    if search:
        queryset = queryset.filter(
            Q(po_number__icontains=search) |
            Q(supplier__name__icontains=search) |
            Q(supplier__code__icontains=search)
        )

    date_range = request.GET.get('date_range')
    if date_range and ' to ' in date_range:
        start_date, end_date = date_range.split(' to ')
        queryset = queryset.filter(order_date__range=[start_date, end_date])

    # Order by latest first
    queryset = queryset.order_by('-order_date', '-id')

    # Pagination
    paginator = Paginator(queryset, 20)  # Show 20 per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Add action flags to each PO for the template
    for po in page_obj:
        po.can_edit = po.status in ['draft', 'pending']
        po.can_cancel = po.status not in ['received', 'cancelled', 'closed']

    # Statistics
    summary_qs = PurchaseOrder.objects.all()
    if company_name:
        summary_qs = summary_qs.filter(company=company_name)

    total_po_count = summary_qs.count()
    total_po_value = summary_qs.aggregate(
        total=Sum('total_amount')
    )['total'] or 0
    pending_count = summary_qs.filter(
        status__in=['draft', 'pending', 'approved']
    ).count()

    # Get all active suppliers for filter dropdown
    suppliers = Supplier.objects.filter(is_active=True)
    if company_name:
        suppliers = suppliers.filter(company=company_name)
    suppliers = suppliers.order_by('name')

    context = {
        'purchase_orders': page_obj,
        'page_obj': page_obj,
        'is_paginated': page_obj.has_other_pages(),
        'suppliers': suppliers,
        'total_po_count': total_po_count,
        'total_po_value': f"{total_po_value:,.0f}",
        'pending_count': pending_count,
        'this_month_count': summary_qs.filter(
            order_date__month=timezone.now().month,
            order_date__year=timezone.now().year
        ).count(),
        'status_choices': PurchaseOrder.STATUS_CHOICES,
    }

    return render(request, 'purchasing/po_list.html', context)


@login_required
def po_detail(request, po_id):
    """
    View purchase order details
    """
    po_qs = PurchaseOrder.objects.select_related('supplier', 'created_by')
    company_name = request.session.get('current_company_name')
    if company_name:
        po_qs = po_qs.filter(company=company_name)
    po = get_object_or_404(po_qs, id=po_id)

    # Get lines, then attach related objects using the temporary item_id/unit_id fields
    lines = list(PurchaseOrderLine.objects.filter(po=po))
    # preload items and units for efficiency
    item_ids = [l.item_id for l in lines if l.item_id]
    unit_ids = [l.unit_id for l in lines if l.unit_id]
    items_map = {i.id: i for i in Item.objects.filter(id__in=item_ids)} if item_ids else {}
    units_map = {u.id: u for u in Unit.objects.filter(id__in=unit_ids)} if unit_ids else {}
    for l in lines:
        l.item_obj = items_map.get(l.item_id)
        l.unit_obj = units_map.get(l.unit_id)

    # Get approvals
    approvals = PurchaseOrderApproval.objects.filter(po=po).order_by('level')

    # Calculate totals
    subtotal = Decimal('0')
    tax_total = Decimal('0')

    for line in lines:
        line_total = line.quantity_ordered * line.unit_price
        line_tax = line_total * (line.tax_rate / 100)
        subtotal += line_total
        tax_total += line_tax

    withholding_tax = _calculate_withholding_tax(subtotal, po.currency)
    grand_total = subtotal + tax_total + po.shipping_cost - po.discount
    net_grand_total = grand_total - withholding_tax

    # Get related receipts if any
    receipts = GoodsReceipt.objects.filter(po=po).select_related('received_by')

    context = {
        'po': po,
        'lines': lines,
        'approvals': approvals,
        'subtotal': subtotal,
        'tax_total': tax_total,
        'withholding_tax': withholding_tax,
        'grand_total': grand_total,
        'net_grand_total': net_grand_total,
        'receipts': receipts,
        'can_edit': po.status in ['draft', 'pending'],
        'can_cancel': po.status not in ['received', 'cancelled', 'closed'],
        'can_receive': po.status in ['approved', 'ordered', 'partial'],
        'can_send': po.status in ['approved'],
        'can_approve': po.status in ['pending'] and not approvals.filter(status='approved').exists(),
    }
    return render(request, 'purchasing/po_detail.html', context)


@login_required
def po_create(request):
    """
    Create a new purchase order
    """
    if request.method == 'POST':
        # Process form submission
        try:
            # Get form data
            supplier_id = request.POST.get('supplier')
            po_number = request.POST.get('po_number')
            order_date = request.POST.get('order_date')
            expected_date = request.POST.get('expected_delivery_date') or request.POST.get('expected_date')
            shipping_method = request.POST.get('shipping_method')
            shipping_address = request.POST.get('shipping_address')
            billing_address = request.POST.get('billing_address')
            notes = request.POST.get('notes')
            terms = request.POST.get('terms')
            shipping_cost = Decimal(request.POST.get('shipping_cost', 0))
            discount = Decimal(request.POST.get('discount', 0))
            currency = request.POST.get('currency', 'USD')
            payment_terms = request.POST.get('payment_terms')
            company_name = request.session.get('current_company_name')
            
            if not company_name:
                messages.error(request, "Please select a company first.")
                return redirect('core:home')

            # Validate required fields
            if not supplier_id or not po_number or not order_date:
                messages.error(request, "Please fill in all required fields.")
                return redirect('purchasing:po_create')

            # Get supplier
            supplier = get_object_or_404(Supplier, id=supplier_id)

            # Create purchase order
            po = PurchaseOrder.objects.create(
                company=company_name,
                po_number=po_number,
                supplier=supplier,
                order_date=order_date,
                expected_delivery_date=expected_date if expected_date else None,
                shipping_method=shipping_method,
                shipping_address=shipping_address,
                billing_address=billing_address,
                notes=notes,
                terms=terms,
                shipping_cost=shipping_cost,
                discount=discount,
                currency=currency,
                payment_terms=payment_terms,
                status='draft',
                created_by=request.user
            )

            # Process order lines
            line_items = request.POST.getlist('item')
            quantities = request.POST.getlist('quantity')
            units = request.POST.getlist('unit')
            unit_prices = request.POST.getlist('unit_price')
            vat_rate = Decimal(request.POST.get('vat_rate_selection', '0'))

            total_amount = Decimal('0')
            withholding_base = Decimal('0')

            for i in range(len(line_items)):
                if line_items[i] and quantities[i] and quantities[i].strip():
                    item_id = line_items[i]
                    if item_id and item_id.strip():
                        try:
                            item = _resolve_po_item(item_id, unit_id=units[i] if i < len(units) else None)
                            quantity = Decimal(quantities[i]) if quantities[i] else Decimal('0')
                            unit_price = Decimal(unit_prices[i]) if i < len(unit_prices) and unit_prices[i] else Decimal('0')

                            unit = None
                            if units[i] and units[i].strip():
                                try:
                                    unit = Unit.objects.get(id=int(units[i]))
                                except (Unit.DoesNotExist, ValueError):
                                    unit = None

                            if quantity > 0:
                                line = PurchaseOrderLine.objects.create(
                                    po=po,
                                    item=f"{item.code} - {item.name}",
                                    item_id=item.id,
                                    unit=str(unit.id) if unit else '',
                                    unit_id=unit.id if unit else None,
                                    quantity_ordered=quantity,
                                    unit_price=unit_price,
                                    tax_rate=vat_rate
                                )

                                line_total = line.quantity_ordered * line.unit_price
                                tax_amount = line_total * (line.tax_rate / 100)
                                total_amount += line_total + tax_amount
                                withholding_base += line_total
                        except (Item.DoesNotExist, ValueError, InvalidOperation) as e:
                            print(f"Error processing line {i}: {e}")
                            continue

            # Update PO total
            withholding_tax = _calculate_withholding_tax(withholding_base, currency)
            po.total_amount = total_amount + shipping_cost - discount - withholding_tax
            po.save()

            messages.success(request, f"Purchase Order {po.po_number} created successfully.")

            # Handle action buttons
            action = request.POST.get('action')
            if action == 'save_and_new':
                return redirect('purchasing:po_create')
            elif action == 'save_and_send':
                # Redirect to send page or send email
                messages.info(request, "PO saved. You can now send it to the supplier.")
                return redirect('purchasing:po_detail', po_id=po.id)
            else:
                return redirect('purchasing:po_detail', po_id=po.id)

        except Exception as e:
            messages.error(request, f"Error creating purchase order: {str(e)}")
            return redirect('purchasing:po_create')

    # GET request - show form
    company_name = request.session.get('current_company_name')
    suppliers = Supplier.objects.filter(is_active=True)
    if company_name:
        suppliers = suppliers.filter(company=company_name)
    suppliers = suppliers.order_by('name')

    # Load purchasable inventory items for dropdown
    items = Item.objects.filter(
        is_active=True,
        is_purchased=True
    ).select_related('unit').order_by('code')

    units = Unit.objects.all().order_by('name')

    # Check if coming from requisition
    requisition_id = request.GET.get('requisition')
    initial_data = {}
    if requisition_id:
        try:
            requisition = PurchaseRequisition.objects.get(id=requisition_id)
            initial_data = {
                'requisition': requisition,
                'expected_date': requisition.required_date,
            }
        except PurchaseRequisition.DoesNotExist:
            pass

    context = {
        'suppliers': suppliers,
        'items': items,
        'units': units,
        'requisition_id': requisition_id,
        'initial_data': initial_data,
        'now': timezone.now(),
    }
    return render(request, 'purchasing/po_form.html', context)


@login_required
def po_edit(request, po_id):
    """
    Edit a purchase order
    """
    po_qs = PurchaseOrder.objects.all()
    company_name = request.session.get('current_company_name')
    if company_name:
        po_qs = po_qs.filter(company=company_name)
    po = get_object_or_404(po_qs, id=po_id)

    # Check if PO can be edited
    if po.status not in ['draft', 'pending']:
        messages.error(request, f"Cannot edit PO with status '{po.status}'.")
        return redirect('purchasing:po_detail', po_id=po.id)

    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Update PO fields
                po.supplier_id = request.POST.get('supplier')
                po.po_number = request.POST.get('po_number')
                po.order_date = request.POST.get('order_date')
                po.expected_delivery_date = request.POST.get('expected_delivery_date') or request.POST.get('expected_date')
                po.shipping_method = request.POST.get('shipping_method')
                po.shipping_address = request.POST.get('shipping_address')
                po.billing_address = request.POST.get('billing_address')
                po.notes = request.POST.get('notes')
                po.terms = request.POST.get('terms')
                po.shipping_cost = Decimal(request.POST.get('shipping_cost', 0))
                po.discount = Decimal(request.POST.get('discount', 0))
                po.currency = request.POST.get('currency', 'USD')
                po.payment_terms = request.POST.get('payment_terms')
                po.save()

                # Handle line items - delete existing and recreate
                po.lines.all().delete()

                line_items = request.POST.getlist('item')
                quantities = request.POST.getlist('quantity')
                units = request.POST.getlist('unit')
                unit_prices = request.POST.getlist('unit_price')
                vat_rate = Decimal(request.POST.get('vat_rate_selection', '0'))

                total_amount = Decimal('0')
                withholding_base = Decimal('0')

                for i in range(len(line_items)):
                    if line_items[i] and quantities[i] and quantities[i].strip():
                        item_id = line_items[i]
                        if item_id and item_id.strip():
                            try:
                                item_obj = _resolve_po_item(item_id, unit_id=units[i] if i < len(units) else None)
                                quantity = Decimal(quantities[i]) if quantities[i] else Decimal('0')
                                unit_price = Decimal(unit_prices[i]) if i < len(unit_prices) and unit_prices[i] else Decimal('0')

                                if quantity > 0:
                                    # build a readable string for the item field
                                    item_str = f"{item_obj.code} - {item_obj.name}"

                                    line = PurchaseOrderLine.objects.create(
                                        po=po,
                                        item=item_str,
                                        item_id=item_obj.id,
                                        unit=str(units[i]) if i < len(units) and units[i] else '',
                                        unit_id=int(units[i]) if i < len(units) and units[i] else None,
                                        quantity_ordered=quantity,
                                        unit_price=unit_price,
                                        tax_rate=vat_rate
                                    )

                                    line_total = line.quantity_ordered * line.unit_price
                                    tax_amount = line_total * (line.tax_rate / 100)
                                    total_amount += line_total + tax_amount
                                    withholding_base += line_total
                            except (ValueError, InvalidOperation) as e:
                                print(f"Error processing line {i}: {e}")
                                continue

                withholding_tax = _calculate_withholding_tax(withholding_base, po.currency)
                po.total_amount = total_amount + po.shipping_cost - po.discount - withholding_tax
                po.save()

                messages.success(request, f"Purchase Order {po.po_number} updated successfully.")
                return redirect('purchasing:po_detail', po_id=po.id)

        except Exception as e:
            messages.error(request, f"Error updating purchase order: {str(e)}")
            return redirect('purchasing:po_edit', po_id=po.id)

    # GET request - show form with existing data
    suppliers = Supplier.objects.filter(is_active=True)
    if company_name:
        suppliers = suppliers.filter(company=company_name)
    suppliers = suppliers.order_by('name')

    # Load items for dropdowns
    items = Item.objects.filter(is_active=True).select_related('unit').order_by('code')

    units = Unit.objects.all().order_by('name')

    # prepare existing order lines with related objects
    order_lines = list(po.lines.all())
    item_ids = [l.item_id for l in order_lines if l.item_id]
    unit_ids = [l.unit_id for l in order_lines if l.unit_id]
    items_map = {i.id: i for i in Item.objects.filter(id__in=item_ids)} if item_ids else {}
    units_map = {u.id: u for u in Unit.objects.filter(id__in=unit_ids)} if unit_ids else {}
    
    # Create a list of dictionaries with the line data instead of trying to set attributes
    lines_data = []
    for l in order_lines:
        item_obj = items_map.get(l.item_id)
        unit_obj = units_map.get(l.unit_id)
        subtotal = l.quantity_ordered * l.unit_price
        
        lines_data.append({
            'id': l.id,
            'item': l.item,
            'item_id': l.item_id,
            'item_obj': item_obj,
            'unit': l.unit,
            'unit_id': l.unit_id,
            'unit_obj': unit_obj,
            'quantity_ordered': l.quantity_ordered,
            'unit_price': l.unit_price,
            'tax_rate': l.tax_rate,
            'subtotal': subtotal,
            'notes': l.notes,
        })

    context = {
        'po': po,
        'suppliers': suppliers,
        'items': items,
        'units': units,
        'order_lines': lines_data,  # Use the list of dictionaries instead of model instances
        'vat_rate': order_lines[0].tax_rate if order_lines else Decimal('15'),
    }
    return render(request, 'purchasing/po_form.html', context)

@login_required
@require_POST
def po_delete(request, po_id):
    """
    Delete a purchase order
    """
    po = get_object_or_404(PurchaseOrder, id=po_id)
    po_number = po.po_number
    po.delete()
    messages.success(request, f"Purchase Order {po_number} deleted successfully.")
    return redirect('purchasing:po_list')


@login_required
@require_POST
def po_send(request, po_id):
    """
    Send purchase order to supplier (change status to 'ordered')
    """
    po = get_object_or_404(PurchaseOrder, id=po_id)
    po.status = 'ordered'
    po.save()
    messages.success(request, f"PO {po.po_number} sent to supplier.")
    # Here you would add email sending logic
    # send_po_email(po)
    return redirect('purchasing:po_detail', po_id=po.id)


@login_required
@require_POST
def receive_po(request, po_id):
    """
    Receive purchase order (create goods receipt and update stock)
    """
    po = get_object_or_404(PurchaseOrder, id=po_id)

    if po.status not in ['approved', 'ordered', 'partial']:
        messages.error(request, f"Cannot receive a PO with status '{po.get_status_display()}'.")
        return redirect('purchasing:po_detail', po_id=po.id)

    try:
        with transaction.atomic():
            received_date = request.POST.get('received_date') or timezone.now().date()
            receipt_number = request.POST.get('receipt_number')
            
            # Auto-generate receipt number if not provided
            if not receipt_number:
                last_receipt = GoodsReceipt.objects.order_by('-id').first()
                if last_receipt and last_receipt.receipt_number and last_receipt.receipt_number.startswith('GRN-'):
                    try:
                        last_num = int(last_receipt.receipt_number.split('-')[-1])
                        new_num = last_num + 1
                    except:
                        new_num = 1
                else:
                    new_num = 1
                today = timezone.now().strftime('%Y%m%d')
                receipt_number = f"GRN-{today}-{new_num:04d}"
            
            notes = request.POST.get('notes', '')

            # Get or create a default receiving warehouse
            receiving_warehouse, _ = Warehouse.objects.get_or_create(
                code='RECV-WH',
                defaults={'name': 'Receiving Warehouse', 'warehouse_type': 'receiving', 'is_active': True}
            )

            # Create goods receipt
            receipt = GoodsReceipt.objects.create(
                receipt_number=receipt_number,
                po=po,
                receipt_date=received_date,
                notes=notes,
                received_by=request.user,
                warehouse=receiving_warehouse.name
            )

            # Update inventory for each line
            all_lines_received = True
            for line in po.lines.all():
                quantity_to_receive = line.remaining
                if quantity_to_receive <= 0:
                    continue

                item = Item.objects.get(id=line.item_id)

                # Create a Lot for this receipt
                lot = Lot.objects.create(
                    item=item,
                    batch_number=f"GRN-{receipt.id}-{item.id}",
                    manufacturing_date=timezone.now().date(),
                    expiry_date=timezone.now().date() + timezone.timedelta(days=item.shelf_life_days or 365),
                    initial_quantity=quantity_to_receive,
                    current_quantity=quantity_to_receive,
                    is_active=True,
                    notes=f"Received from PO {po.po_number}"
                )

                # Create a stock transaction
                StockTransaction.objects.create(
                    transaction_type='receipt',
                    item=item,
                    lot=lot,
                    warehouse_to=receiving_warehouse,
                    quantity=quantity_to_receive,
                    unit_cost=line.unit_price,
                    transaction_date=receipt.receipt_date,
                    reference=f"GRN-{receipt.receipt_number}",
                )

                # Create receipt line
                GoodsReceiptLine.objects.create(
                    receipt=receipt,
                    po_line=line,
                    quantity_received=quantity_to_receive,
                    lot=lot.batch_number
                )

                # Update the item's main stock count
                item.current_stock = (item.current_stock or 0) + quantity_to_receive
                item.save(update_fields=['current_stock', 'updated_at'])

                # Update the PO line's received quantity
                line.quantity_received = (line.quantity_received or 0) + quantity_to_receive
                line.save(update_fields=['quantity_received'])

                if line.remaining > 0:
                    all_lines_received = False

            # Update PO status
            if all_lines_received:
                po.status = 'received'
            else:
                po.status = 'partial'
            po.received_date = received_date
            po.save()

            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'receipt_id': receipt.id, 'redirect_url': reverse('purchasing:po_detail', args=[po.id])})

            messages.success(request, f"PO {po.po_number} received successfully. Receipt #{receipt.receipt_number} created. Inventory updated.")
            return redirect('purchasing:po_detail', po_id=po.id)

    except Exception as e:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
        messages.error(request, f"Error receiving PO: {str(e)}")
        return redirect('purchasing:po_detail', po_id=po.id)


@login_required
@require_POST
def po_close(request, po_id):
    """
    Close a purchase order
    """
    po = get_object_or_404(PurchaseOrder, id=po_id)
    po.status = 'closed'
    po.save()
    messages.success(request, f"PO {po.po_number} closed.")
    return redirect('purchasing:po_detail', po_id=po.id)


@login_required
@require_POST
def po_cancel(request, po_id):
    """
    Cancel a purchase order
    """
    po = get_object_or_404(PurchaseOrder, id=po_id)
    reason = request.POST.get('reason')
    notes = request.POST.get('notes')
    po.status = 'cancelled'
    po.notes = f"Cancelled: {reason}\n{notes}" if notes else f"Cancelled: {reason}"
    po.save()
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': True, 'redirect_url': reverse('purchasing:po_detail', args=[po.id])})
    messages.success(request, f"PO {po.po_number} cancelled.")
    return redirect('purchasing:po_detail', po_id=po.id)


@login_required
def po_print(request, po_id):
    """
    Print purchase order (return PDF or print-friendly HTML)
    """
    po_qs = PurchaseOrder.objects.select_related('supplier', 'created_by')
    company_name = request.session.get('current_company_name')
    if company_name:
        po_qs = po_qs.filter(company=company_name)
    po = get_object_or_404(po_qs, id=po_id)

    lines = list(PurchaseOrderLine.objects.filter(po=po))
    unit_ids = [l.unit_id for l in lines if l.unit_id]
    units_map = {u.id: u for u in Unit.objects.filter(id__in=unit_ids)} if unit_ids else {}
    for l in lines:
        l.unit_obj = units_map.get(l.unit_id)

    # Calculate totals
    subtotal = Decimal('0')
    tax_total = Decimal('0')

    for line in lines:
        line_total = line.quantity_ordered * line.unit_price
        line_tax = line_total * (line.tax_rate / 100)
        subtotal += line_total
        tax_total += line_tax

    withholding_tax = _calculate_withholding_tax(subtotal, po.currency)
    grand_total = subtotal + tax_total + po.shipping_cost - po.discount
    net_grand_total = grand_total - withholding_tax

    company_obj = None
    company_id = request.session.get('current_company_id')
    if company_id:
        company_obj = Company.objects.filter(id=company_id, is_active=True).first()

    context = {
        'po': po,
        'lines': lines,
        'subtotal': subtotal,
        'tax_total': tax_total,
        'withholding_tax': withholding_tax,
        'grand_total': grand_total,
        'net_grand_total': net_grand_total,
        'company_name': company_obj.name if company_obj else (request.session.get('current_company_name') or po.company or 'Company'),
        'company_address': company_obj.address if company_obj else '',
        'company_phone': company_obj.phone if company_obj else '',
        'company_email': company_obj.email if company_obj else '',
        'company_tax_id': company_obj.tin_number if company_obj else '',
    }
    return render(request, 'purchasing/po_print.html', context)


def download_po_pdf(request, po_id):
    po = get_object_or_404(PurchaseOrder, id=po_id)
    pdf_buffer = generate_po_pdf(po)
    return FileResponse(pdf_buffer, as_attachment=True, filename=f'{po.po_number}.pdf')


@login_required
def export_pos(request):
    """
    Export purchase orders to CSV/Excel
    """
    import csv
    from django.http import HttpResponse

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="purchase_orders.csv"'

    writer = csv.writer(response)
    writer.writerow(['PO Number', 'Supplier', 'Order Date', 'Expected Date', 'Received Date', 'Total Amount', 'Status', 'Created By', 'Created Date'])

    pos = PurchaseOrder.objects.all().select_related('supplier', 'created_by')
    for po in pos:
        writer.writerow([
            po.po_number,
            po.supplier.name if po.supplier else '',
            po.order_date,
            po.expected_delivery_date,
            po.received_date,
            f"{po.total_amount:.2f}",
            po.status,
            po.created_by.username if po.created_by else '',
            po.created_at,
        ])

    return response


# ============================
# Purchase Order Approval Views
# ============================

@login_required
@require_POST
def po_approve(request, po_id, level):
    """
    Approve purchase order at specific approval level
    """
    po = get_object_or_404(PurchaseOrder, id=po_id)
    
    # Check if approval level exists
    approval, created = PurchaseOrderApproval.objects.get_or_create(
        po=po,
        level=level,
        defaults={
            'approver': request.user,
            'status': 'approved',
            'approved_at': timezone.now()
        }
    )
    
    if not created:
        approval.status = 'approved'
        approval.approver = request.user
        approval.approved_at = timezone.now()
        approval.save()
    
    # Check if all required approvals are done
    required_levels = [1, 2]  # You can make this configurable
    approved_levels = po.approvals.filter(status='approved').count()
    
    if approved_levels >= len(required_levels):
        po.status = 'approved'
        po.approved_by = request.user
        po.save()
    
    messages.success(request, f"PO {po.po_number} approved at level {level}.")
    return redirect('purchasing:po_detail', po_id=po.id)


@login_required
@require_POST
def po_reject(request, po_id, level):
    """
    Reject purchase order at specific approval level
    """
    po = get_object_or_404(PurchaseOrder, id=po_id)
    comment = request.POST.get('comment', '')
    
    approval, created = PurchaseOrderApproval.objects.get_or_create(
        po=po,
        level=level,
        defaults={
            'approver': request.user,
            'status': 'rejected',
            'comment': comment
        }
    )
    
    if not created:
        approval.status = 'rejected'
        approval.comment = comment
        approval.save()
    
    po.status = 'cancelled'
    po.save()
    
    messages.warning(request, f"PO {po.po_number} rejected at level {level}.")
    return redirect('purchasing:po_detail', po_id=po.id)


# ============================
# Supplier Views
# ============================

@login_required
def supplier_list(request):
    """
    List all suppliers with filtering
    """
    # Base queryset
    suppliers = Supplier.objects.all().order_by('name')

    # Apply filters
    search = request.GET.get('search')
    if search:
        suppliers = suppliers.filter(
            Q(name__icontains=search) |
            Q(code__icontains=search) |
            Q(email__icontains=search) |
            Q(phone__icontains=search) |
            Q(contact_person__icontains=search) |
            Q(tax_id__icontains=search)
        )

    is_active = request.GET.get('is_active')
    if is_active == 'true':
        suppliers = suppliers.filter(is_active=True)
    elif is_active == 'false':
        suppliers = suppliers.filter(is_active=False)

    # Pagination
    paginator = Paginator(suppliers, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Statistics
    total_count = Supplier.objects.count()
    active_count = Supplier.objects.filter(is_active=True).count()
    inactive_count = total_count - active_count

    context = {
        'suppliers': page_obj,
        'page_obj': page_obj,
        'is_paginated': page_obj.has_other_pages(),
        'total_count': total_count,
        'active_count': active_count,
        'inactive_count': inactive_count,
        'export_url': reverse('purchasing:export_suppliers') + f'?{request.GET.urlencode()}',
    }

    return render(request, 'purchasing/supplier_list.html', context)


@login_required
def supplier_detail(request, supplier_id):
    """
    Enhanced supplier detail view with comprehensive metrics and real data
    """
    company_name = request.session.get('current_company_name')
    
    # Get supplier with related data
    supplier = get_object_or_404(Supplier, id=supplier_id)
    
    # Base queryset for purchase orders (respecting company context)
    po_queryset = PurchaseOrder.objects.filter(supplier=supplier)
    if company_name:
        po_queryset = po_queryset.filter(company=company_name)
    
    # Recent purchase orders (with related data for efficiency)
    recent_pos = po_queryset.select_related(
        'created_by'
    ).order_by('-order_date')[:10]

    # Comprehensive supplier performance metrics
    total_pos = po_queryset.count()
    
    # Total spend (received POs)
    total_spend = po_queryset.filter(
        status='received'
    ).aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
    
    # Outstanding amount (ordered but not received)
    outstanding = po_queryset.filter(
        status__in=['approved', 'ordered', 'partial']
    ).aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
    
    # Get on-time delivery rate
    received_pos = po_queryset.filter(
        status='received',
        received_date__isnull=False,
        expected_delivery_date__isnull=False
    )
    
    on_time_count = received_pos.filter(
        received_date__lte=F('expected_delivery_date')
    ).count()
    
    on_time_rate = 0
    if received_pos.count() > 0:
        on_time_rate = (on_time_count / received_pos.count()) * 100
    
    # Calculate average lead time (from order to receipt)
    lead_times = []
    for po in received_pos:
        if po.received_date and po.order_date:
            lead_days = (po.received_date - po.order_date).days
            lead_times.append(lead_days)
    
    avg_lead_time = round(sum(lead_times) / len(lead_times), 1) if lead_times else 0
    
    # Calculate this month's spend
    today = timezone.now().date()
    month_start = today.replace(day=1)
    monthly_spend = po_queryset.filter(
        status='received',
        received_date__gte=month_start,
        received_date__lte=today
    ).aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
    
    # Get pending approvals count for this supplier's POs
    pending_approvals_count = PurchaseOrderApproval.objects.filter(
        po__supplier=supplier,
        status='pending'
    ).count()
    
    # Get recent goods receipts for this supplier
    recent_receipts = GoodsReceipt.objects.filter(
        po__supplier=supplier
    ).select_related('po', 'received_by').order_by('-receipt_date')[:5]
    
    # Get recent bills if PurchaseBill model exists (optional)
    recent_bills = []
    try:
        from apps.accounting.models import PurchaseBill
        recent_bills = PurchaseBill.objects.filter(
            supplier=supplier
        ).order_by('-bill_date')[:5]
    except (ImportError, AttributeError):
        pass  # PurchaseBill model doesn't exist, skip
    
    # Get recent payments if Payment model exists (optional)
    recent_payments = []
    try:
        from apps.accounting.models import Payment
        recent_payments = Payment.objects.filter(
            supplier=supplier
        ).order_by('-date')[:5]
    except (ImportError, AttributeError):
        pass  # Payment model doesn't exist, skip
    
    # Get top purchased items from this supplier
    top_items = PurchaseOrderLine.objects.filter(
        po__supplier=supplier,
        po__status='received'
    ).values('item').annotate(
        total_quantity=Sum('quantity_received'),
        total_spent=Sum(F('quantity_received') * F('unit_price'))
    ).order_by('-total_spent')[:5]
    
    # Get item details for top items
    for item_data in top_items:
        try:
            item = Item.objects.get(id=item_data['item'])
            item_data['item_code'] = item.code
            item_data['item_name'] = item.name
            item_data['unit'] = item.unit.abbreviation if item.unit else ''
        except Item.DoesNotExist:
            item_data['item_code'] = 'Unknown'
            item_data['item_name'] = 'Unknown'
    
    # Calculate performance rating based on actual data if not set manually
    if supplier.performance_rating == 0 and received_pos.count() > 5:
        # Calculate weighted score (60% on-time, 40% quality if available)
        delivery_score = on_time_rate
        quality_score = float(supplier.quality_rating or 0) * 20  # Convert 0-5 to 0-100
        calculated_rating = (delivery_score * 0.6 + quality_score * 0.4) / 20  # Convert back to 0-5 scale
        display_rating = round(calculated_rating, 1)
    else:
        display_rating = float(supplier.performance_rating or 0)
    
    context = {
        # Basic supplier info
        'supplier': supplier,
        
        # Recent transactions
        'recent_pos': recent_pos,
        'recent_receipts': recent_receipts,
        'recent_bills': recent_bills,
        'recent_payments': recent_payments,
        
        # Performance metrics
        'po_count': total_pos,
        'total_spend': total_spend,
        'monthly_spend': monthly_spend,
        'outstanding': outstanding,
        'on_time_rate': round(on_time_rate, 1),
        'avg_lead_time': avg_lead_time,
        
        # Ratings
        'performance_rating': display_rating,
        'quality_rating': float(supplier.quality_rating or 0),
        
        # Additional metrics
        'pending_approvals_count': pending_approvals_count,
        'top_items': top_items,
        
        # Helper for template
        'has_receipts': recent_receipts.exists(),
        'has_bills': len(recent_bills) > 0,
        'has_payments': len(recent_payments) > 0,
    }
    
    return render(request, 'purchasing/supplier_detail.html', context)


@login_required
def supplier_create(request):
    """
    Create a new supplier
    """
    if request.method == 'POST':
        try:
            company_name = request.session.get('current_company_name')
            
            supplier = Supplier.objects.create(
                company=company_name,
                code=request.POST.get('code'),
                name=request.POST.get('name'),
                contact_person=request.POST.get('contact_person'),
                email=request.POST.get('email'),
                phone=request.POST.get('phone'),
                address=request.POST.get('address'),
                payment_terms_days=int(request.POST.get('payment_terms_days', 30)),
                tax_id=request.POST.get('tax_id'),
                is_active=request.POST.get('is_active') == 'on'
            )
            messages.success(request, f"Supplier {supplier.name} created successfully.")
            return redirect('purchasing:supplier_detail', supplier_id=supplier.id)
        except Exception as e:
            messages.error(request, f"Error creating supplier: {str(e)}")
            return redirect('purchasing:supplier_create')

    return render(request, 'purchasing/supplier_form.html')


@login_required
def supplier_edit(request, supplier_id):
    """
    Edit an existing supplier
    """
    supplier = get_object_or_404(Supplier, id=supplier_id)

    if request.method == 'POST':
        try:
            # Update supplier fields
            supplier.code = request.POST.get('code')
            supplier.name = request.POST.get('name')
            supplier.contact_person = request.POST.get('contact_person')
            supplier.email = request.POST.get('email')
            supplier.phone = request.POST.get('phone')
            supplier.address = request.POST.get('address')
            supplier.payment_terms_days = int(request.POST.get('payment_terms_days', 30))
            supplier.tax_id = request.POST.get('tax_id')
            supplier.is_active = request.POST.get('is_active') == 'on'
            supplier.save()

            messages.success(request, f'Supplier {supplier.name} updated successfully.')
            return redirect('purchasing:supplier_detail', supplier_id=supplier.id)

        except Exception as e:
            messages.error(request, f'Error updating supplier: {str(e)}')
            return redirect('purchasing:supplier_edit', supplier_id=supplier.id)

    context = {
        'supplier': supplier,
    }
    return render(request, 'purchasing/supplier_form.html', context)


@login_required
@require_POST
def supplier_delete(request, supplier_id):
    """
    Delete a supplier
    """
    supplier = get_object_or_404(Supplier, id=supplier_id)
    supplier_name = supplier.name
    supplier.delete()
    messages.success(request, f'Supplier {supplier_name} deleted successfully.')
    return redirect('purchasing:supplier_list')


@login_required
def supplier_performance(request, supplier_id):
    """
    View supplier performance metrics
    """
    supplier = get_object_or_404(Supplier, id=supplier_id)

    # Get purchase orders for this supplier
    purchase_orders = PurchaseOrder.objects.filter(supplier=supplier).order_by('-order_date')

    # Calculate core metrics
    total_orders = purchase_orders.count()
    total_spent = purchase_orders.filter(status='received').aggregate(
        total=Sum('total_amount')
    )['total'] or Decimal('0.00')

    received_orders = purchase_orders.filter(
        status='received',
        received_date__isnull=False
    )
    delivered_orders_count = received_orders.count()

    on_time_orders = received_orders.filter(
        expected_delivery_date__isnull=False,
        received_date__lte=F('expected_delivery_date')
    ).count()
    late_orders = max(delivered_orders_count - on_time_orders, 0)

    on_time_rate = (on_time_orders / delivered_orders_count * 100) if delivered_orders_count else 0

    lead_time_values = []
    for po in received_orders.exclude(order_date__isnull=True):
        if po.received_date and po.order_date:
            lead_time_values.append((po.received_date - po.order_date).days)
    avg_lead_time = round(sum(lead_time_values) / len(lead_time_values), 1) if lead_time_values else 0

    # Build recent order performance rows
    recent_orders = []
    for po in purchase_orders[:20]:
        lead_days = None
        if po.received_date and po.order_date:
            lead_days = (po.received_date - po.order_date).days

        delay_days = None
        is_on_time = None
        if po.received_date and po.expected_delivery_date:
            delay_days = (po.received_date - po.expected_delivery_date).days
            is_on_time = delay_days <= 0

        recent_orders.append({
            'po': po,
            'lead_days': lead_days,
            'delay_days': delay_days,
            'is_on_time': is_on_time,
        })

    # Weighted score (60% delivery + 40% quality)
    quality_rating = float(supplier.quality_rating or 0)  # expected out of 5
    delivery_score = on_time_rate
    quality_score = (quality_rating / 5.0) * 100 if quality_rating > 0 else 0
    score = round((delivery_score * 0.6) + (quality_score * 0.4), 1)

    if score >= 90:
        score_grade = 'A'
    elif score >= 80:
        score_grade = 'B'
    elif score >= 70:
        score_grade = 'C'
    else:
        score_grade = 'D'

    context = {
        'supplier': supplier,
        'recent_orders': recent_orders,
        'total_orders': total_orders,
        'delivered_orders_count': delivered_orders_count,
        'on_time_orders': on_time_orders,
        'late_orders': late_orders,
        'total_spent': total_spent,
        'on_time_rate': round(on_time_rate, 1),
        'avg_lead_time': avg_lead_time,
        'score': score,
        'score_grade': score_grade,
        'current_date': timezone.now(),
    }

    return render(request, 'purchasing/supplier_performance.html', context)


@login_required
def export_suppliers(request):
    """
    Export suppliers to CSV/Excel
    """
    import csv
    from django.http import HttpResponse

    suppliers = Supplier.objects.all().order_by('name')

    # Apply same filters as supplier_list, but never paginate
    search = (request.GET.get('search') or '').strip()
    if search:
        suppliers = suppliers.filter(
            Q(name__icontains=search) |
            Q(code__icontains=search) |
            Q(email__icontains=search) |
            Q(phone__icontains=search) |
            Q(contact_person__icontains=search)
        )

    is_active = request.GET.get('is_active')
    if is_active == 'true':
        suppliers = suppliers.filter(is_active=True)
    elif is_active == 'false':
        suppliers = suppliers.filter(is_active=False)

    export_format = (request.GET.get('format') or 'csv').lower()

    if export_format == 'excel':
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            export_format = 'csv'
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Suppliers"

            # Header row
            headers = ['Code', 'Name', 'Contact Person', 'Email', 'Phone',
                      'Address', 'City', 'Country', 'Payment Terms', 'Tax ID', 'Active']
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')

            # Data rows
            for row_num, supplier in enumerate(suppliers, 2):
                ws.cell(row=row_num, column=1).value = supplier.code
                ws.cell(row=row_num, column=2).value = supplier.name
                ws.cell(row=row_num, column=3).value = supplier.contact_person
                ws.cell(row=row_num, column=4).value = supplier.email
                ws.cell(row=row_num, column=5).value = supplier.phone
                ws.cell(row=row_num, column=6).value = supplier.address
                ws.cell(row=row_num, column=7).value = supplier.city
                ws.cell(row=row_num, column=8).value = supplier.country
                ws.cell(row=row_num, column=9).value = supplier.payment_terms_days
                ws.cell(row=row_num, column=10).value = supplier.tax_id
                ws.cell(row=row_num, column=11).value = 'Yes' if supplier.is_active else 'No'

            # Auto-adjust column widths
            for col_num in range(1, len(headers) + 1):
                column_letter = get_column_letter(col_num)
                ws.column_dimensions[column_letter].width = 15

            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="suppliers.xlsx"'
            wb.save(response)
            return response

    # Default CSV export
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="suppliers.csv"'

    writer = csv.writer(response)
    writer.writerow(['Code', 'Name', 'Contact Person', 'Email', 'Phone', 'Address', 'Payment Terms', 'Tax ID', 'Active'])

    for supplier in suppliers:
        writer.writerow([
            supplier.code,
            supplier.name,
            supplier.contact_person,
            supplier.email,
            supplier.phone,
            supplier.address,
            supplier.payment_terms_days,
            supplier.tax_id,
            'Yes' if supplier.is_active else 'No',
        ])

    return response


# ============================
# Requisition Views
# ============================

@login_required
def requisition_list(request):
    """
    List all purchase requisitions
    """
    company_name = request.session.get('current_company_name')
    requisitions = PurchaseRequisition.objects.all().select_related('requested_by')
    if company_name:
        requisitions = requisitions.filter(company=company_name)
    requisitions = requisitions.order_by('-requested_date')
    
    # Apply filters
    status = request.GET.get('status')
    if status:
        requisitions = requisitions.filter(status=status)
    
    context = {
        'requisitions': requisitions,
        'status_choices': PurchaseRequisition.STATUS_CHOICES,
    }
    return render(request, 'purchasing/requisition_list.html', context)


@login_required
def requisition_detail(request, requisition_id):
    """
    View requisition details
    """
    company_name = request.session.get('current_company_name')
    requisition_qs = PurchaseRequisition.objects.select_related('requested_by')
    if company_name:
        requisition_qs = requisition_qs.filter(company=company_name)
    requisition = get_object_or_404(requisition_qs, id=requisition_id)
    
    lines = requisition.lines.all().select_related('item', 'unit')
    
    context = {
        'requisition': requisition,
        'lines': lines,
        'total_estimate': requisition.lines.aggregate(total=Sum(F('quantity') * F('unit_price_estimate')))['total'] or 0,
    }
    return render(request, 'purchasing/requisition_detail.html', context)


@login_required
def requisition_create(request):
    """
    Create a new purchase requisition
    """
    if request.method == 'POST':
        try:
            company_name = request.session.get('current_company_name')
            
            requisition = PurchaseRequisition.objects.create(
                company=company_name,
                requisition_number=request.POST.get('requisition_number'),
                required_date=request.POST.get('required_date'),
                requested_by=request.user,
                notes=request.POST.get('notes'),
                status='draft'
            )

            # Process requisition lines
            items = request.POST.getlist('item')
            quantities = request.POST.getlist('quantity')
            units = request.POST.getlist('unit')
            unit_prices = request.POST.getlist('unit_price_estimate')

            for i in range(len(items)):
                if items[i] and quantities[i] and quantities[i].strip():
                    item = Item.objects.get(id=int(items[i]))
                    unit = Unit.objects.get(id=int(units[i])) if units[i] else None
                    
                    PurchaseRequisitionLine.objects.create(
                        requisition=requisition,
                        item=item,
                        quantity=Decimal(quantities[i]),
                        unit=unit,
                        unit_price_estimate=Decimal(unit_prices[i]) if i < len(unit_prices) and unit_prices[i] else Decimal('0')
                    )

            messages.success(request, f"Requisition {requisition.requisition_number} created successfully.")
            return redirect('purchasing:requisition_detail', requisition_id=requisition.id)
        except Exception as e:
            messages.error(request, f"Error creating requisition: {str(e)}")
            return redirect('purchasing:requisition_create')

    # GET request - show form
    items = Item.objects.filter(is_active=True, is_purchased=True).select_related('unit').order_by('code')
    units = Unit.objects.all().order_by('name')
    
    # Generate requisition number
    last_req = PurchaseRequisition.objects.order_by('-id').first()
    if last_req and last_req.requisition_number and last_req.requisition_number.startswith('REQ-'):
        try:
            last_num = int(last_req.requisition_number.split('-')[-1])
            new_num = last_num + 1
        except:
            new_num = 1
    else:
        new_num = 1
    
    today = timezone.now().strftime('%Y%m%d')
    requisition_number = f"REQ-{today}-{new_num:04d}"

    context = {
        'items': items,
        'units': units,
        'requisition_number': requisition_number,
        'today': timezone.now().date(),
    }
    return render(request, 'purchasing/requisition_form.html', context)


@login_required
def requisition_edit(request, requisition_id):
    """
    Edit requisition
    """
    company_name = request.session.get('current_company_name')
    requisition_qs = PurchaseRequisition.objects.all()
    if company_name:
        requisition_qs = requisition_qs.filter(company=company_name)
    requisition = get_object_or_404(requisition_qs, id=requisition_id)
    
    if requisition.status not in ['draft', 'rejected']:
        messages.error(request, "Cannot edit requisition that has been submitted or approved.")
        return redirect('purchasing:requisition_detail', requisition_id=requisition.id)

    if request.method == 'POST':
        try:
            requisition.required_date = request.POST.get('required_date')
            requisition.notes = request.POST.get('notes')
            requisition.save()

            # Delete existing lines and recreate
            requisition.lines.all().delete()

            items = request.POST.getlist('item')
            quantities = request.POST.getlist('quantity')
            units = request.POST.getlist('unit')
            unit_prices = request.POST.getlist('unit_price_estimate')

            for i in range(len(items)):
                if items[i] and quantities[i] and quantities[i].strip():
                    item = Item.objects.get(id=int(items[i]))
                    unit = Unit.objects.get(id=int(units[i])) if units[i] else None
                    
                    PurchaseRequisitionLine.objects.create(
                        requisition=requisition,
                        item=item,
                        quantity=Decimal(quantities[i]),
                        unit=unit,
                        unit_price_estimate=Decimal(unit_prices[i]) if i < len(unit_prices) and unit_prices[i] else Decimal('0')
                    )

            messages.success(request, "Requisition updated successfully.")
            return redirect('purchasing:requisition_detail', requisition_id=requisition.id)
        except Exception as e:
            messages.error(request, f"Error updating requisition: {str(e)}")
            return redirect('purchasing:requisition_edit', requisition_id=requisition.id)

    # GET request - show form with existing data
    items = Item.objects.filter(is_active=True, is_purchased=True).select_related('unit').order_by('code')
    units = Unit.objects.all().order_by('name')
    lines = requisition.lines.all().select_related('item', 'unit')

    context = {
        'requisition': requisition,
        'items': items,
        'units': units,
        'lines': lines,
    }
    return render(request, 'purchasing/requisition_form.html', context)


@login_required
@require_POST
def requisition_delete(request, requisition_id):
    """
    Delete requisition
    """
    company_name = request.session.get('current_company_name')
    requisition_qs = PurchaseRequisition.objects.all()
    if company_name:
        requisition_qs = requisition_qs.filter(company=company_name)
    requisition = get_object_or_404(requisition_qs, id=requisition_id)
    
    if requisition.status not in ['draft', 'rejected']:
        messages.error(request, "Cannot delete requisition that has been submitted or approved.")
        return redirect('purchasing:requisition_detail', requisition_id=requisition.id)
    
    req_number = requisition.requisition_number
    requisition.delete()
    messages.success(request, f"Requisition {req_number} deleted.")
    return redirect('purchasing:requisition_list')


@login_required
@require_POST
def requisition_submit(request, requisition_id):
    """
    Submit requisition for approval
    """
    company_name = request.session.get('current_company_name')
    requisition_qs = PurchaseRequisition.objects.all()
    if company_name:
        requisition_qs = requisition_qs.filter(company=company_name)
    requisition = get_object_or_404(requisition_qs, id=requisition_id)
    
    if requisition.status != 'draft':
        messages.error(request, "Only draft requisitions can be submitted.")
        return redirect('purchasing:requisition_detail', requisition_id=requisition.id)
    
    requisition.status = 'submitted'
    requisition.save()
    messages.success(request, "Requisition submitted for approval.")
    return redirect('purchasing:requisition_detail', requisition_id=requisition.id)


@login_required
@require_POST
def requisition_approve(request, requisition_id):
    """
    Approve requisition
    """
    company_name = request.session.get('current_company_name')
    requisition_qs = PurchaseRequisition.objects.all()
    if company_name:
        requisition_qs = requisition_qs.filter(company=company_name)
    requisition = get_object_or_404(requisition_qs, id=requisition_id)
    
    if requisition.status != 'submitted':
        messages.error(request, "Only submitted requisitions can be approved.")
        return redirect('purchasing:requisition_detail', requisition_id=requisition.id)
    
    requisition.status = 'approved'
    requisition.save()
    messages.success(request, "Requisition approved.")
    return redirect('purchasing:requisition_detail', requisition_id=requisition.id)


@login_required
def requisition_reject(request, requisition_id):
    """
    Reject requisition
    """
    company_name = request.session.get('current_company_name')
    requisition_qs = PurchaseRequisition.objects.all()
    if company_name:
        requisition_qs = requisition_qs.filter(company=company_name)
    requisition = get_object_or_404(requisition_qs, id=requisition_id)

    if request.method == 'POST':
        reason = request.POST.get('reason')
        requisition.status = 'rejected'
        requisition.notes = f"Rejected: {reason}" if reason else "Rejected"
        requisition.save()
        messages.success(request, "Requisition rejected.")
        return redirect('purchasing:requisition_detail', requisition_id=requisition.id)

    context = {'requisition': requisition}
    return render(request, 'purchasing/requisition_reject.html', context)


@login_required
def create_po_from_requisition(request, requisition_id):
    """
    Create purchase order from requisition
    """
    requisition = get_object_or_404(PurchaseRequisition, id=requisition_id)
    # Redirect to PO create with pre-filled data
    return redirect(f"{reverse('purchasing:po_create')}?requisition={requisition_id}")


# ============================
# Goods Receipt Views
# ============================

@login_required
def goods_receipt_list(request):
    """
    List all goods receipts
    """
    company_name = request.session.get('current_company_name')
    receipts = GoodsReceipt.objects.all().select_related(
        'po', 'received_by'
    ).order_by('-receipt_date')
    if company_name:
        receipts = receipts.filter(company=company_name)
    
    context = {'receipts': receipts}
    return render(request, 'purchasing/receipt_list.html', context)


@login_required
def goods_receipt_detail(request, receipt_id):
    """
    View goods receipt details
    """
    company_name = request.session.get('current_company_name')
    receipt_qs = GoodsReceipt.objects.select_related('po', 'received_by')
    if company_name:
        receipt_qs = receipt_qs.filter(company=company_name)
    receipt = get_object_or_404(receipt_qs, id=receipt_id)
    
    lines = GoodsReceiptLine.objects.filter(receipt=receipt).select_related('po_line')
    
    # Calculate total value
    total_value = sum(line.line_total for line in lines)
    
    context = {
        'receipt': receipt, 
        'lines': lines,
        'total_value': total_value,
    }
    return render(request, 'purchasing/receipt_detail.html', context)


@login_required
def goods_receipt_create(request):
    """
    Create a new goods receipt
    """
    if request.method == 'POST':
        try:
            po_id = request.POST.get('po')
            po = get_object_or_404(PurchaseOrder, id=po_id)

            if po.status not in ['approved', 'ordered', 'partial']:
                messages.error(request, f"Cannot receive a PO with status '{po.get_status_display()}'.")
                return redirect('purchasing:goods_receipt_create', po_id=po.id)

            company_name = request.session.get('current_company_name')

            with transaction.atomic():
                received_date = request.POST.get('receipt_date') or timezone.now().date()
                receipt = GoodsReceipt.objects.create(
                    receipt_number=request.POST.get('receipt_number'),
                    po=po,
                    receipt_date=received_date,
                    warehouse=request.POST.get('warehouse', ''),
                    notes=request.POST.get('notes'),
                    received_by=request.user,
                    company=company_name
                )

                # Get or create a default receiving warehouse
                receiving_warehouse, _ = Warehouse.objects.get_or_create(
                    code='RECV-WH',
                    defaults={'name': 'Receiving Warehouse', 'warehouse_type': 'receiving', 'is_active': True}
                )

                # Create receipt lines for each PO line
                something_received = False
                for po_line in po.lines.all():
                    qty_str = request.POST.get(f'qty_{po_line.id}', '0').strip()
                    if not qty_str:
                        continue

                    try:
                        quantity_received = Decimal(qty_str)
                    except InvalidOperation:
                        continue

                    if quantity_received > 0:
                        something_received = True
                        if quantity_received > po_line.remaining:
                            raise ValidationError(f"Cannot receive {quantity_received} for item {po_line.item}. Only {po_line.remaining} remaining.")

                        item = Item.objects.get(id=po_line.item_id)

                        # Create a Lot for this receipt
                        lot = Lot.objects.create(
                            item=item,
                            batch_number=f"GRN-{receipt.id}-{item.id}",
                            manufacturing_date=timezone.now().date(),
                            expiry_date=timezone.now().date() + timezone.timedelta(days=item.shelf_life_days or 365),
                            initial_quantity=quantity_received,
                            current_quantity=quantity_received,
                            is_active=True,
                            notes=f"Received from PO {po.po_number}"
                        )

                        # Create a stock transaction
                        StockTransaction.objects.create(
                            transaction_type='receipt',
                            item=item,
                            lot=lot,
                            warehouse_to=receiving_warehouse,
                            quantity=quantity_received,
                            unit_cost=po_line.unit_price,
                            transaction_date=receipt.receipt_date,
                            reference=f"GRN-{receipt.receipt_number}",
                        )

                        GoodsReceiptLine.objects.create(
                            receipt=receipt,
                            po_line=po_line,
                            quantity_received=quantity_received,
                            lot=lot.batch_number,
                            notes=request.POST.get(f'notes_{po_line.id}', '')
                        )

                        # Update the item's main stock count
                        item.current_stock = (item.current_stock or 0) + quantity_received
                        item.save(update_fields=['current_stock', 'updated_at'])

                        # Update the PO line's received quantity
                        po_line.quantity_received = (po_line.quantity_received or 0) + quantity_received
                        po_line.save(update_fields=['quantity_received'])

                if not something_received:
                    raise ValidationError("No quantities were entered to be received.")

                # Update PO status
                all_lines_received = all(line.remaining <= 0 for line in po.lines.all())
                if all_lines_received:
                    po.status = 'received'
                else:
                    po.status = 'partial'
                po.received_date = received_date
                po.save(update_fields=['status', 'received_date'])

                messages.success(request, f"Goods receipt {receipt.receipt_number} created successfully.")
                return redirect('purchasing:goods_receipt_detail', receipt_id=receipt.id)
        except (ValidationError, Exception) as e:
            messages.error(request, f"Error creating goods receipt: {str(e)}")
            po_id = request.POST.get('po')
            if po_id:
                return redirect(f"{reverse('purchasing:goods_receipt_create')}?po_id={po_id}")
            return redirect('purchasing:goods_receipt_create')

    # GET request - show form
    po_id = request.GET.get('po_id')
    initial_po = None
    po_lines_data = []
    
    if po_id:
        initial_po = get_object_or_404(PurchaseOrder, id=po_id)
        # Get lines with remaining quantities
        for line in initial_po.lines.all():
            if line.remaining > 0:
                po_lines_data.append({
                    'id': line.id,
                    'item_name': line.item,
                    'quantity_ordered': line.quantity_ordered,
                    'quantity_received': line.quantity_received,
                    'remaining': line.remaining,
                    'unit': line.unit,
                })

    company_name = request.session.get('current_company_name')
    pos = PurchaseOrder.objects.filter(status__in=['ordered', 'approved', 'partial'])
    if company_name:
        pos = pos.filter(company=company_name)
    pos = pos.order_by('-order_date')
    
    warehouses = ['Main Warehouse', 'Secondary Warehouse', 'Receiving Dock']
    
    # Generate receipt number automatically
    last_receipt = GoodsReceipt.objects.order_by('-id').first()
    if last_receipt and last_receipt.receipt_number and last_receipt.receipt_number.startswith('GRN-'):
        try:
            last_num = int(last_receipt.receipt_number.split('-')[-1])
            new_num = last_num + 1
        except:
            new_num = 1
    else:
        new_num = 1
    
    today = timezone.now().strftime('%Y%m%d')
    receipt_number = f"GRN-{today}-{new_num:04d}"
    
    context = {
        'pos': pos,
        'initial_po': initial_po,
        'po_lines': po_lines_data,
        'warehouses': warehouses,
        'receipt_number': receipt_number,
        'today': timezone.now().date(),
    }
    return render(request, 'purchasing/receipt_form.html', context)

@login_required
def goods_receipt_edit(request, receipt_id):
    """
    Edit goods receipt
    """
    company_name = request.session.get('current_company_name')
    receipt_qs = GoodsReceipt.objects.all()
    if company_name:
        receipt_qs = receipt_qs.filter(company=company_name)
    receipt = get_object_or_404(receipt_qs, id=receipt_id)

    if request.method == 'POST':
        try:
            receipt.receipt_date = request.POST.get('receipt_date')
            receipt.warehouse = request.POST.get('warehouse', '')
            receipt.notes = request.POST.get('notes')
            receipt.save()

            messages.success(request, "Goods receipt updated successfully.")
            return redirect('purchasing:goods_receipt_detail', receipt_id=receipt.id)
        except Exception as e:
            messages.error(request, f"Error updating goods receipt: {str(e)}")
            return redirect('purchasing:goods_receipt_edit', receipt_id=receipt.id)

    warehouses = ['Main Warehouse', 'Secondary Warehouse', 'Receiving Dock']
    
    context = {
        'receipt': receipt,
        'warehouses': warehouses,
    }
    return render(request, 'purchasing/receipt_form.html', context)


@login_required
@require_POST
def goods_receipt_delete(request, receipt_id):
    """
    Delete goods receipt
    """
    company_name = request.session.get('current_company_name')
    receipt_qs = GoodsReceipt.objects.all()
    if company_name:
        receipt_qs = receipt_qs.filter(company=company_name)
    receipt = get_object_or_404(receipt_qs, id=receipt_id)
    
    receipt_number = receipt.receipt_number
    receipt.delete()
    messages.success(request, f"Goods receipt {receipt_number} deleted.")
    return redirect('purchasing:goods_receipt_list')


@login_required
def export_receipts(request):
    """
    Export goods receipts to CSV/Excel
    """
    import csv
    from django.http import HttpResponse

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="goods_receipts.csv"'

    writer = csv.writer(response)
    writer.writerow(['Receipt Number', 'PO Number', 'Received Date', 'Warehouse', 'Notes', 'Created By', 'Created Date'])

    receipts = GoodsReceipt.objects.all().select_related('po', 'received_by')
    for receipt in receipts:
        writer.writerow([
            receipt.receipt_number,
            receipt.po.po_number if receipt.po else '',
            receipt.receipt_date,
            receipt.warehouse,
            receipt.notes,
            receipt.received_by.username if receipt.received_by else '',
            receipt.created_at,
        ])

    return response


# ============================
# Report Views
# ============================

@login_required
def purchasing_report(request):
    """
    Main purchasing report
    """
    company_name = request.session.get('current_company_name')
    
    # Get date range from request
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    if not date_from:
        date_from = (timezone.now() - timedelta(days=30)).strftime('%Y-%m-%d')
    if not date_to:
        date_to = timezone.now().strftime('%Y-%m-%d')
    
    # Base querysets
    po_qs = PurchaseOrder.objects.all()
    if company_name:
        po_qs = po_qs.filter(company=company_name)
    
    # Get statistics
    total_pos = po_qs.filter(
        order_date__range=[date_from, date_to]
    ).count()
    
    total_spend = po_qs.filter(
        order_date__range=[date_from, date_to],
        status='received'
    ).aggregate(total=Sum('total_amount'))['total'] or 0
    
    pending_pos = po_qs.filter(
        status__in=['draft', 'pending', 'approved']
    ).count()
    
    # Get top suppliers
    top_suppliers = Supplier.objects.annotate(
        total_spent=Sum('purchase_orders__total_amount', 
                       filter=Q(purchase_orders__status='received',
                                purchase_orders__order_date__range=[date_from, date_to]))
    ).order_by('-total_spent')[:10]
    
    # Get monthly trend
    monthly_data = []
    for i in range(6):
        month = timezone.now() - timedelta(days=30*i)
        month_start = month.replace(day=1)
        if i == 0:
            month_end = timezone.now()
        else:
            next_month = month_start + timedelta(days=32)
            month_end = next_month.replace(day=1) - timedelta(days=1)
        
        month_spend = po_qs.filter(
            status='received',
            received_date__range=[month_start, month_end]
        ).aggregate(total=Sum('total_amount'))['total'] or 0
        
        monthly_data.append({
            'month': month_start.strftime('%b %Y'),
            'spend': month_spend,
        })
    
    context = {
        'page_title': "Purchasing Reports",
        'date_from': date_from,
        'date_to': date_to,
        'total_pos': total_pos,
        'total_spend': total_spend,
        'pending_pos': pending_pos,
        'top_suppliers': top_suppliers,
        'monthly_data': monthly_data,
    }
    return render(request, 'purchasing/report.html', context)


@login_required
def spend_analysis(request):
    """
    Spend analysis report
    """
    company_name = request.session.get('current_company_name')
    
    # Get date range
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    if not date_from:
        date_from = (timezone.now() - timedelta(days=365)).strftime('%Y-%m-%d')
    if not date_to:
        date_to = timezone.now().strftime('%Y-%m-%d')
    
    # Base queryset
    po_qs = PurchaseOrder.objects.filter(
        status='received',
        received_date__range=[date_from, date_to]
    )
    if company_name:
        po_qs = po_qs.filter(company=company_name)
    
    # Get spend by supplier
    supplier_spend = po_qs.values(
        'supplier__name', 'supplier__code'
    ).annotate(
        total_spend=Sum('total_amount'),
        order_count=Count('id')
    ).order_by('-total_spend')[:20]
    
    # Get spend by month
    monthly_spend = po_qs.extra(
        select={'month': "strftime('%%Y-%%m', received_date)"}
    ).values('month').annotate(
        total=Sum('total_amount')
    ).order_by('month')
    
    # Get spend by status
    status_spend = PurchaseOrder.objects.values('status').annotate(
        total=Sum('total_amount'),
        count=Count('id')
    ).order_by('status')
    
    total_spend = po_qs.aggregate(total=Sum('total_amount'))['total'] or 0
    
    context = {
        'supplier_spend': supplier_spend,
        'monthly_spend': monthly_spend,
        'status_spend': status_spend,
        'total_spend': total_spend,
        'date_from': date_from,
        'date_to': date_to,
    }
    return render(request, 'purchasing/spend_analysis.html', context)


@login_required
def lead_time_report(request):
    """
    Supplier lead time report
    """
    company_name = request.session.get('current_company_name')
    
    # Calculate lead times for received POs
    pos = PurchaseOrder.objects.filter(
        status='received',
        received_date__isnull=False,
        order_date__isnull=False
    ).select_related('supplier')
    
    if company_name:
        pos = pos.filter(company=company_name)

    lead_times = []
    supplier_stats = {}
    
    for po in pos:
        if po.received_date and po.order_date:
            lead_time = (po.received_date - po.order_date).days
            
            lead_times.append({
                'supplier': po.supplier.name,
                'supplier_id': po.supplier.id,
                'po_number': po.po_number,
                'order_date': po.order_date,
                'received_date': po.received_date,
                'lead_time': lead_time,
            })
            
            # Aggregate by supplier
            if po.supplier.id not in supplier_stats:
                supplier_stats[po.supplier.id] = {
                    'name': po.supplier.name,
                    'total_lead_time': 0,
                    'count': 0,
                    'min': 999,
                    'max': 0
                }
            
            stats = supplier_stats[po.supplier.id]
            stats['total_lead_time'] += lead_time
            stats['count'] += 1
            stats['min'] = min(stats['min'], lead_time)
            stats['max'] = max(stats['max'], lead_time)
    
    # Calculate averages
    for stats in supplier_stats.values():
        stats['avg'] = round(stats['total_lead_time'] / stats['count'], 1)
    
    context = {
        'lead_times': lead_times,
        'supplier_stats': supplier_stats.values(),
    }
    return render(request, 'purchasing/lead_time_report.html', context)


@login_required
def po_status_report(request):
    """
    Purchase order status report
    """
    company_name = request.session.get('current_company_name')
    
    po_qs = PurchaseOrder.objects.all()
    if company_name:
        po_qs = po_qs.filter(company=company_name)
    
    status_counts = po_qs.values('status').annotate(
        count=Count('id'),
        total_value=Sum('total_amount')
    ).order_by('status')
    
    # Get counts by supplier
    supplier_counts = po_qs.values('supplier__name').annotate(
        count=Count('id')
    ).order_by('-count')[:15]
    
    context = {
        'status_counts': status_counts,
        'supplier_counts': supplier_counts,
    }
    return render(request, 'purchasing/po_status_report.html', context)


# ============================
# AJAX Views
# ============================

@login_required
def ajax_supplier_info(request, supplier_id):
    """
    AJAX endpoint to get supplier information
    """
    try:
        supplier = Supplier.objects.get(id=supplier_id)
        data = {
            'id': supplier.id,
            'name': supplier.name,
            'code': supplier.code,
            'contact_person': supplier.contact_person,
            'email': supplier.email,
            'phone': supplier.phone,
            'address': supplier.address,
            'payment_terms': supplier.payment_terms_days,
            'payment_terms_days': supplier.payment_terms_days,
            'vat_registered': bool(supplier.tax_id or supplier.tin),
            'tax_id': supplier.tax_id,
        }
        return JsonResponse(data)
    except Supplier.DoesNotExist:
        return JsonResponse({'error': 'Supplier not found'}, status=404)


@login_required
def ajax_po_lines(request, po_id):
    """
    AJAX endpoint to get PO lines
    """
    try:
        po = PurchaseOrder.objects.get(id=po_id)
        lines = []
        for line in po.lines.all():
            lines.append({
                'id': line.id,
                'item': line.item,
                'item_id': line.item_id,
                'quantity_ordered': float(line.quantity_ordered),
                'quantity_received': float(line.quantity_received),
                'unit': line.unit,
                'unit_price': float(line.unit_price),
                'tax_rate': float(line.tax_rate),
                'total_price': float(line.total_price),
                'remaining': float(line.remaining),
            })
        return JsonResponse({'lines': lines})
    except PurchaseOrder.DoesNotExist:
        return JsonResponse({'error': 'PO not found'}, status=404)


@login_required
def ajax_item_price(request, item_id, supplier_id):
    """
    AJAX endpoint to get item price from supplier
    """
    try:
        # Get item
        item = Item.objects.get(id=item_id)
        
        # You could implement supplier-specific pricing here
        # For now, return the standard unit cost
        data = {
            'item_id': item.id,
            'unit_price': float(item.unit_cost or 0),
            'unit': item.unit.abbreviation if item.unit else '',
            'tax_rate': 15,  # Default VAT rate
        }
        return JsonResponse(data)
    except Item.DoesNotExist:
        return JsonResponse({'error': 'Item not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def ajax_check_po_number(request):
    """Check if PO number already exists"""
    po_number = request.GET.get('po_number')
    if po_number:
        exists = PurchaseOrder.objects.filter(po_number=po_number).exists()
        return JsonResponse({'exists': exists})
    return JsonResponse({'exists': False})


@login_required
def ajax_check_receipt_number(request):
    """Check if receipt number already exists"""
    receipt_number = request.GET.get('receipt_number')
    if receipt_number:
        exists = GoodsReceipt.objects.filter(receipt_number=receipt_number).exists()
        return JsonResponse({'exists': exists})
    return JsonResponse({'exists': False})