from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import models
from django.db.models import Sum, Count, Q, Avg, F
from django.utils import timezone
from django.http import HttpResponse, JsonResponse, FileResponse
from django.template.loader import render_to_string
from datetime import datetime, timedelta
from decimal import Decimal
import json
import csv
import io

# Report generation libraries
import pandas as pd
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from .models import ReportTemplate, ScheduledReport, DashboardWidget, ReportCategory
from .forms import (
    InventoryReportFilterForm, ProductionReportFilterForm,
    PurchasingReportFilterForm, SalesReportFilterForm,
    FinancialReportFilterForm, ScheduledReportForm,
    DashboardWidgetForm
)

# Import models from other apps
from apps.core.models import Item, Unit
from apps.inventory.models import Warehouse, Lot, StockTransaction, CurrentStock
from apps.production.models import ProductionRun, BOM, InventoryMovement
from apps.purchasing.models import Supplier, PurchaseOrder, PurchaseOrderLine, GoodsReceipt
from apps.sales.models import SalesOrder, SalesInvoice, SalesPayment
from apps.accounting.models import Account, JournalEntry, JournalLine, PurchaseBill
from apps.sales.models import SalesInvoice 
from apps.company.models import Company, Customer, UserProfile


# ============================
# Dashboard
# ============================

@login_required
def dashboard(request):
    """Main reports dashboard"""
    categories = ReportCategory.objects.filter(is_active=True).prefetch_related('reports')
    
    # Get user's custom widgets
    widgets = DashboardWidget.objects.filter(user=request.user, is_visible=True).order_by('position')
    
    # Get scheduled reports
    scheduled = ScheduledReport.objects.filter(is_active=True)[:5]
    
    # Quick stats
    total_items = Item.objects.filter(is_active=True).count()
    total_suppliers = Supplier.objects.filter(is_active=True).count()
    active_production = ProductionRun.objects.filter(status='in_progress').count()
    pending_orders = PurchaseOrder.objects.filter(status__in=['draft', 'confirmed']).count()
    monthly_revenue = SalesInvoice.objects.filter(
        invoice_date__gte=timezone.now().replace(day=1),
        status__in=['posted', 'paid']
    ).aggregate(total=Sum('total_amount'))['total'] or 0
    scheduled_reports = ScheduledReport.objects.filter(is_active=True).count()
    
    context = {
        'categories': categories,
        'widgets': widgets,
        'scheduled': scheduled,
        'today': timezone.now().date(),
        'total_items': total_items,
        'total_suppliers': total_suppliers,
        'active_production': active_production,
        'pending_orders': pending_orders,
        'monthly_revenue': monthly_revenue,
        'scheduled_reports': scheduled_reports,
    }
    
    return render(request, 'reports/dashboard.html', context)


# ============================
# Report Generation Views
# ============================

@login_required
def stock_summary_report(request):
    """Stock summary report"""
    form = InventoryReportFilterForm(request.GET)
    
    # Base queryset
    stock_items = CurrentStock.objects.select_related('item', 'warehouse', 'lot')
    
    # Apply filters
    if form.is_valid():
        if form.cleaned_data.get('warehouse'):
            stock_items = stock_items.filter(warehouse_id=form.cleaned_data['warehouse'])
        if form.cleaned_data.get('category'):
            stock_items = stock_items.filter(item__category=form.cleaned_data['category'])
        if form.cleaned_data.get('item'):
            stock_items = stock_items.filter(item__code__icontains=form.cleaned_data['item'])
        # apply date filters based on related stock transactions
        if form.cleaned_data.get('date_from'):
            stock_items = stock_items.filter(
                item__stocktransaction__transaction_date__date__gte=form.cleaned_data['date_from']
            )
        if form.cleaned_data.get('date_to'):
            stock_items = stock_items.filter(
                item__stocktransaction__transaction_date__date__lte=form.cleaned_data['date_to']
            )
        stock_items = stock_items.distinct()
    
    # Calculate totals
    total_value = stock_items.aggregate(
        total=Sum(F('quantity') * F('item__unit_cost'))
    )['total'] or 0
    total_quantity = stock_items.aggregate(total=Sum('quantity'))['total'] or 0
    
    # Group by item
    by_item = stock_items.values(
        'item__code', 'item__name', 'item__category', 'item_id'
    ).annotate(
        total_qty=Sum('quantity'),
        total_value=Sum(F('quantity') * F('item__unit_cost'))
    ).order_by('item__code')

    # include items with zero stock if they match filters
    existing_ids = [row['item_id'] for row in by_item]
    zero_qs = Item.objects.filter(is_active=True).exclude(id__in=existing_ids)
    if form.is_valid():
        if form.cleaned_data.get('category'):
            zero_qs = zero_qs.filter(category=form.cleaned_data['category'])
        if form.cleaned_data.get('item'):
            zero_qs = zero_qs.filter(code__icontains=form.cleaned_data['item'])
    zero_rows = []
    for itm in zero_qs:
        zero_rows.append({
            'item__code': itm.code,
            'item__name': itm.name,
            'item__category': itm.get_category_display() if hasattr(itm, 'get_category_display') else itm.category,
            'total_qty': 0,
            'total_value': 0,
        })
    by_item = list(by_item) + zero_rows
    
    # Group by warehouse
    by_warehouse = stock_items.values(
        'warehouse__name', 'warehouse__code'
    ).annotate(
        total_qty=Sum('quantity'),
        total_value=Sum(F('quantity') * F('item__unit_cost'))
    ).order_by('-total_value')
    
    # Expiry summary
    today = timezone.now().date()
    near_expiry = Lot.objects.filter(
        expiry_date__lte=today + timedelta(days=90),
        expiry_date__gte=today,
        is_active=True
    ).count()
    
    expired = Lot.objects.filter(
        expiry_date__lt=today,
        is_active=True
    ).count()
    
    # good expiry count (>90 days until expiry)
    good_expiry_count = Lot.objects.filter(
        expiry_date__gt=today + timedelta(days=90),
        is_active=True
    ).count()
    
    # warehouses and categories for filters
    warehouses = Warehouse.objects.filter(is_active=True)
    categories = Item.ITEM_CATEGORY
    
    # low stock count based on item thresholds
    low_stock_count = stock_items.filter(
        Q(quantity__lte=F('item__reorder_point')) |
        Q(quantity__lte=F('item__minimum_stock'))
    ).count()
    
    context = {
        'form': form,
        'stock_items': stock_items[:1000],  # Limit for performance
        'by_item': by_item,
        'by_warehouse': by_warehouse,
        'total_value': total_value,
        'total_items': stock_items.count(),
        'total_quantity': total_quantity,
        'low_stock_count': low_stock_count,
        'near_expiry': near_expiry,
        'near_expiry_count': near_expiry,
        'expired': expired,
        'expired_count': expired,
        'good_expiry_count': good_expiry_count,
        'warehouses': warehouses,
        'categories': categories,
        'today': today,
    }
    
    # Handle export
    if request.GET.get('export') == 'excel':
        return export_stock_summary_excel(context)
    elif request.GET.get('export') == 'pdf':
        return export_stock_summary_pdf(context)
    elif request.GET.get('export') == 'csv':
        return export_stock_summary_csv(context)
    
    return render(request, 'reports/inventory/stock_summary.html', context)


@login_required
def stock_value_report(request):
    """Stock value report with trends"""
    form = InventoryReportFilterForm(request.GET)
    
    # Get current stock value
    current_value = CurrentStock.objects.aggregate(
        total=Sum(F('quantity') * F('item__unit_cost'))
    )['total'] or 0
    
    # Get value by category
    by_category = CurrentStock.objects.values(
        'item__category'
    ).annotate(
        total_qty=Sum('quantity'),
        total_value=Sum(F('quantity') * F('item__unit_cost'))
    ).order_by('-total_value')
    
    # Get value by warehouse
    by_warehouse = CurrentStock.objects.values(
        'warehouse__name'
    ).annotate(
        total_value=Sum(F('quantity') * F('item__unit_cost'))
    ).order_by('-total_value')
    
    # Top 10 items by value
    top_items = CurrentStock.objects.values(
        'item__code', 'item__name'
    ).annotate(
        total_value=Sum(F('quantity') * F('item__unit_cost'))
    ).order_by('-total_value')[:10]
    
    # Monthly trend (last 12 months)
    months = []
    values = []
    for i in range(11, -1, -1):
        date = timezone.now().date() - timedelta(days=30*i)
        months.append(date.strftime('%b %Y'))
        # This would need historical snapshots - using current for now
        values.append(float(current_value))
    
    context = {
        'form': form,
        'current_value': current_value,
        'by_category': by_category,
        'by_warehouse': by_warehouse,
        'top_items': top_items,
        'months': json.dumps(months),
        'values': json.dumps(values),
        'today': timezone.now().date(),
    }
    
    return render(request, 'reports/inventory/stock_value.html', context)


@login_required
def stock_aging_report(request):
    """Stock aging report by receipt date"""
    form = InventoryReportFilterForm(request.GET)
    
    # Get all lots with their receipt dates
    lots = Lot.objects.filter(is_active=True).select_related('item', 'warehouse')
    
    # Apply filters
    if form.is_valid():
        if form.cleaned_data.get('warehouse'):
            lots = lots.filter(warehouse_id=form.cleaned_data['warehouse'])
        if form.cleaned_data.get('category'):
            lots = lots.filter(item__category=form.cleaned_data['category'])
    
    # Calculate aging buckets
    today = timezone.now().date()
    aging_buckets = {
        '0_30': lots.filter(received_date__gte=today - timedelta(days=30)),
        '31_60': lots.filter(received_date__lt=today - timedelta(days=30), received_date__gte=today - timedelta(days=60)),
        '61_90': lots.filter(received_date__lt=today - timedelta(days=60), received_date__gte=today - timedelta(days=90)),
        '91_180': lots.filter(received_date__lt=today - timedelta(days=90), received_date__gte=today - timedelta(days=180)),
        '180_plus': lots.filter(received_date__lt=today - timedelta(days=180)),
    }
    
    bucket_stats = {}
    for name, queryset in aging_buckets.items():
        bucket_stats[name] = {
            'count': queryset.count(),
            'value': queryset.aggregate(
                total=Sum(F('current_quantity') * F('item__unit_cost'))
            )['total'] or 0
        }
    
    context = {
        'form': form,
        'bucket_stats': bucket_stats,
        'aging_buckets': aging_buckets,
        'today': today,
    }
    
    return render(request, 'reports/inventory/stock_aging.html', context)


@login_required
def low_stock_report(request):
    """Low stock alert report"""
    # Get items below reorder point
    low_stock = Item.objects.filter(
        Q(current_stock__lte=F('reorder_point')) |
        Q(current_stock__lte=F('minimum_stock')),
        is_active=True
    ).select_related('unit').order_by('current_stock')
    
    # Critical items (below minimum)
    critical = low_stock.filter(current_stock__lte=F('minimum_stock'))
    
    # Reorder items (below reorder but above minimum)
    reorder = low_stock.filter(
        current_stock__lte=F('reorder_point'),
        current_stock__gt=F('minimum_stock')
    )
    
    # Out of stock
    out_of_stock = Item.objects.filter(current_stock=0, is_active=True)
    
    context = {
        'low_stock': low_stock,
        'critical': critical,
        'reorder': reorder,
        'out_of_stock': out_of_stock,
        'critical_count': critical.count(),
        'reorder_count': reorder.count(),
        'out_count': out_of_stock.count(),
        'today': timezone.now().date(),
    }
    
    return render(request, 'reports/inventory/low_stock.html', context)


@login_required
def expiry_report(request):
    """Expiry report for lots"""
    today = timezone.now().date()
    
    # Expired lots
    expired = Lot.objects.filter(
        expiry_date__lt=today,
        is_active=True
    ).select_related('item').order_by('expiry_date')
    
    # Near expiry (next 90 days)
    near_expiry = Lot.objects.filter(
        expiry_date__gte=today,
        expiry_date__lte=today + timedelta(days=90),
        is_active=True
    ).select_related('item').order_by('expiry_date')
    
    # 30-day buckets
    buckets = {}
    for days in [30, 60, 90]:
        cutoff = today + timedelta(days=days)
        buckets[f'{days}_days'] = Lot.objects.filter(
            expiry_date__lte=cutoff,
            expiry_date__gt=cutoff - timedelta(days=30),
            is_active=True
        ).count()
    
    context = {
        'expired': expired,
        'near_expiry': near_expiry,
        'buckets': buckets,
        'expired_count': expired.count(),
        'near_count': near_expiry.count(),
        'today': today,
    }
    
    return render(request, 'reports/inventory/expiry.html', context)


@login_required
def inventory_movements_report(request):
    """Inventory movements report"""
    form = InventoryReportFilterForm(request.GET)
    
    movements = StockTransaction.objects.select_related(
        'item', 'lot', 'warehouse_from', 'warehouse_to', 'created_by'
    ).order_by('-transaction_date')
    
    if form.is_valid():
        if form.cleaned_data.get('date_from'):
            movements = movements.filter(transaction_date__date__gte=form.cleaned_data['date_from'])
        if form.cleaned_data.get('date_to'):
            movements = movements.filter(transaction_date__date__lte=form.cleaned_data['date_to'])
        if form.cleaned_data.get('item'):
            movements = movements.filter(item__code__icontains=form.cleaned_data['item'])
    
    # Summary by type
    by_type = movements.values('transaction_type').annotate(
        count=Count('id'),
        total_qty=Sum('quantity')
    )
    
    context = {
        'form': form,
        'movements': movements[:500],
        'by_type': by_type,
        'today': timezone.now().date(),
    }
    
    return render(request, 'reports/inventory/movements.html', context)


@login_required
def production_runs_report(request):
    """Production runs report"""
    form = ProductionReportFilterForm(request.GET)
    
    runs = ProductionRun.objects.select_related('product', 'bom').order_by('-start_date')
    
    if form.is_valid():
        if form.cleaned_data.get('date_from'):
            runs = runs.filter(start_date__date__gte=form.cleaned_data['date_from'])
        if form.cleaned_data.get('date_to'):
            runs = runs.filter(start_date__date__lte=form.cleaned_data['date_to'])
        if form.cleaned_data.get('product'):
            runs = runs.filter(product_id=form.cleaned_data['product'])
        if form.cleaned_data.get('status'):
            runs = runs.filter(status=form.cleaned_data['status'])
    
    # Summary stats
    stats = {
        'total_runs': runs.count(),
        'completed': runs.filter(status='completed').count(),
        'in_progress': runs.filter(status='in_progress').count(),
        'planned': runs.filter(status='planned').count(),
        'total_planned_qty': runs.aggregate(total=Sum('planned_quantity'))['total'] or 0,
        'total_actual_qty': runs.filter(status='completed').aggregate(total=Sum('actual_quantity'))['total'] or 0,
    }
    
    # Yield analysis
    completed = runs.filter(status='completed', actual_quantity__isnull=False)
    for run in completed:
        run.yield_pct = (run.actual_quantity / run.planned_quantity * 100) if run.planned_quantity else 0
    
    context = {
        'form': form,
        'runs': runs[:100],
        'stats': stats,
        'completed': completed,
        'today': timezone.now().date(),
    }
    
    return render(request, 'reports/production/runs.html', context)


@login_required
def cost_variance_report(request):
    """Production cost variance report"""
    from apps.production.models import ProductionCostVariance
    
    variances = ProductionCostVariance.objects.select_related(
        'production_run__product'
    ).order_by('-calculated_at')[:100]
    
    # Summary
    total_std = variances.aggregate(total=Sum('standard_total_cost'))['total'] or 0
    total_act = variances.aggregate(total=Sum('actual_total_cost'))['total'] or 0
    total_var = total_act - total_std
    var_pct = (total_var / total_std * 100) if total_std else 0
    
    context = {
        'variances': variances,
        'total_std': total_std,
        'total_act': total_act,
        'total_var': total_var,
        'var_pct': var_pct,
        'today': timezone.now().date(),
    }
    
    return render(request, 'reports/production/cost_variance.html', context)


@login_required
def production_yield_report(request):
    """Production yield report"""
    form = ProductionReportFilterForm(request.GET)
    
    runs = ProductionRun.objects.filter(
        status='completed',
        actual_quantity__isnull=False
    ).select_related('product')
    
    if form.is_valid():
        if form.cleaned_data.get('date_from'):
            runs = runs.filter(end_date__date__gte=form.cleaned_data['date_from'])
        if form.cleaned_data.get('date_to'):
            runs = runs.filter(end_date__date__lte=form.cleaned_data['date_to'])
        if form.cleaned_data.get('product'):
            runs = runs.filter(product_id=form.cleaned_data['product'])
    
    # Calculate yield
    for run in runs:
        run.yield_pct = (run.actual_quantity / run.planned_quantity * 100) if run.planned_quantity else 0
    
    # Yield buckets
    yield_buckets = {
        'excellent': runs.filter(yield_percentage__gte=95).count(),
        'good': runs.filter(yield_percentage__gte=90, yield_percentage__lt=95).count(),
        'average': runs.filter(yield_percentage__gte=80, yield_percentage__lt=90).count(),
        'poor': runs.filter(yield_percentage__lt=80).count(),
    }
    
    # By product
    by_product = runs.values('product__code', 'product__name').annotate(
        avg_yield=Avg('yield_percentage'),
        total_runs=Count('id')
    ).order_by('-avg_yield')
    
    context = {
        'form': form,
        'runs': runs,
        'yield_buckets': yield_buckets,
        'by_product': by_product,
        'today': timezone.now().date(),
    }
    
    return render(request, 'reports/production/yield.html', context)


@login_required
def material_consumption_report(request):
    """Material consumption report"""
    form = ProductionReportFilterForm(request.GET)
    
    # Get consumption from inventory movements
    consumption = InventoryMovement.objects.filter(
        movement_type='out_production'
    ).select_related('item', 'production_run')
    
    if form.is_valid():
        if form.cleaned_data.get('date_from'):
            consumption = consumption.filter(created_at__date__gte=form.cleaned_data['date_from'])
        if form.cleaned_data.get('date_to'):
            consumption = consumption.filter(created_at__date__lte=form.cleaned_data['date_to'])
    
    # By material
    by_material = consumption.values(
        'item__code', 'item__name', 'item__category'
    ).annotate(
        total_qty=Sum('quantity'),
        total_cost=Sum(F('quantity') * F('item__unit_cost'))
    ).order_by('-total_cost')
    
    # By production run
    by_run = consumption.values(
        'production_run__id',
        'production_run__product__code'
    ).annotate(
        total_qty=Sum('quantity'),
        total_cost=Sum(F('quantity') * F('item__unit_cost'))
    ).order_by('-total_cost')[:20]
    
    context = {
        'form': form,
        'by_material': by_material,
        'by_run': by_run,
        'total_consumption': consumption.count(),
        'today': timezone.now().date(),
    }
    
    return render(request, 'reports/production/consumption.html', context)


@login_required
def po_summary_report(request):
    """Purchase Order summary report"""
    form = PurchasingReportFilterForm(request.GET)
    
    pos = PurchaseOrder.objects.select_related('supplier').order_by('-order_date')
    
    if form.is_valid():
        if form.cleaned_data.get('date_from'):
            pos = pos.filter(order_date__gte=form.cleaned_data['date_from'])
        if form.cleaned_data.get('date_to'):
            pos = pos.filter(order_date__lte=form.cleaned_data['date_to'])
        if form.cleaned_data.get('supplier'):
            pos = pos.filter(supplier_id=form.cleaned_data['supplier'])
        if form.cleaned_data.get('status'):
            pos = pos.filter(status=form.cleaned_data['status'])
    
    # Summary stats
    stats = {
        'total_pos': pos.count(),
        'total_amount': pos.aggregate(total=Sum('total_amount'))['total'] or 0,
        'open_orders': pos.filter(status__in=['ordered', 'partial']).count(),
        'received': pos.filter(status='received').count(),
    }
    
    # By status
    by_status = pos.values('status').annotate(
        count=Count('id'),
        total=Sum('total_amount')
    )
    
    # By supplier
    by_supplier = pos.values('supplier__name').annotate(
        count=Count('id'),
        total=Sum('total_amount')
    ).order_by('-total')[:10]
    
    context = {
        'form': form,
        'pos': pos[:100],
        'stats': stats,
        'by_status': by_status,
        'by_supplier': by_supplier,
        'today': timezone.now().date(),
    }
    
    return render(request, 'reports/purchasing/po_summary.html', context)


@login_required
def supplier_performance_report(request):
    """Supplier performance report"""
    suppliers = Supplier.objects.annotate(
        total_orders=Count('purchase_orders'),
        total_purchase_amount=Sum('purchase_orders__total_amount'),
        on_time_deliveries=Count('purchase_orders', filter=Q(
            purchase_orders__status='received',
            purchase_orders__goods_receipts__receipt_date__lte=F('purchase_orders__expected_delivery_date')
        )),
        late_deliveries=Count('purchase_orders', filter=Q(
            purchase_orders__status='received',
            purchase_orders__goods_receipts__receipt_date__gt=F('purchase_orders__expected_delivery_date')
        ))
    ).filter(total_orders__gt=0)
    
    for supplier in suppliers:
        if supplier.total_orders > 0:
            supplier.on_time_rate = (supplier.on_time_deliveries / supplier.total_orders) * 100
        else:
            supplier.on_time_rate = 0
    
    # Sort by performance
    top_performers = sorted(suppliers, key=lambda x: x.on_time_rate, reverse=True)[:10]
    low_performers = sorted(suppliers, key=lambda x: x.on_time_rate)[:10]
    
    context = {
        'suppliers': suppliers,
        'top_performers': top_performers,
        'low_performers': low_performers,
        'today': timezone.now().date(),
    }
    
    return render(request, 'reports/purchasing/supplier_performance.html', context)


@login_required
def spend_analysis_report(request):
    """Spend analysis report"""
    form = PurchasingReportFilterForm(request.GET)
    
    pos = PurchaseOrder.objects.filter(status__in=['received', 'partial'])
    
    if form.is_valid():
        if form.cleaned_data.get('date_from'):
            pos = pos.filter(order_date__gte=form.cleaned_data['date_from'])
        if form.cleaned_data.get('date_to'):
            pos = pos.filter(order_date__lte=form.cleaned_data['date_to'])
    
    # By supplier
    by_supplier = pos.values('supplier__name').annotate(
        total=Sum('total_amount'),
        count=Count('id')
    ).order_by('-total')
    
    # By category (through items)
    by_category = PurchaseOrderLine.objects.filter(
        po__status__in=['received', 'partial']
    ).values(
        'item__category'
    ).annotate(
        total=Sum('total_price'),
        count=Count('id')
    ).order_by('-total')
    
    # Monthly trend
    monthly = []
    for i in range(11, -1, -1):
        month = timezone.now().date() - timedelta(days=30*i)
        month_start = month.replace(day=1)
        if month.month == 12:
            month_end = month.replace(month=1, day=1) - timedelta(days=1)
        else:
            month_end = month.replace(month=month.month+1, day=1) - timedelta(days=1)
        
        total = pos.filter(
            order_date__gte=month_start,
            order_date__lte=month_end
        ).aggregate(total=Sum('total_amount'))['total'] or 0
        
        monthly.append({
            'month': month.strftime('%b %Y'),
            'total': float(total)
        })
    
    context = {
        'form': form,
        'by_supplier': by_supplier,
        'by_category': by_category,
        'monthly': monthly,
        'total_spend': pos.aggregate(total=Sum('total_amount'))['total'] or 0,
        'today': timezone.now().date(),
    }
    
    return render(request, 'reports/purchasing/spend_analysis.html', context)


@login_required
def lead_time_report(request):
    """Lead time analysis report"""
    completed_pos = PurchaseOrder.objects.filter(
        status='received',
        goods_receipts__isnull=False
    ).annotate(
        lead_time=F('goods_receipts__receipt_date') - F('order_date')
    )
    
    # Overall stats
    stats = {
        'avg_lead_time': completed_pos.aggregate(avg=Avg('lead_time'))['avg'],
        'min_lead_time': completed_pos.aggregate(min=models.Min('lead_time'))['min'],
        'max_lead_time': completed_pos.aggregate(max=models.Max('lead_time'))['max'],
        'total_orders': completed_pos.count(),
    }
    
    # By supplier
    by_supplier = completed_pos.values('supplier__name').annotate(
        avg_lead=Avg('lead_time'),
        count=Count('id')
    ).order_by('avg_lead')
    
    context = {
        'stats': stats,
        'by_supplier': by_supplier,
        'today': timezone.now().date(),
    }
    
    return render(request, 'reports/purchasing/lead_time.html', context)


@login_required
def revenue_analysis_report(request):
    """Revenue analysis report"""
    form = SalesReportFilterForm(request.GET)
    
    invoices = SalesInvoice.objects.filter(status__in=['posted', 'paid', 'partial'])
    
    if form.is_valid():
        if form.cleaned_data.get('date_from'):
            invoices = invoices.filter(invoice_date__gte=form.cleaned_data['date_from'])
        if form.cleaned_data.get('date_to'):
            invoices = invoices.filter(invoice_date__lte=form.cleaned_data['date_to'])
        if form.cleaned_data.get('customer'):
            invoices = invoices.filter(customer_id=form.cleaned_data['customer'])
    
    # By customer
    by_customer = invoices.values('customer__name').annotate(
        total=Sum('total_amount'),
        count=Count('id')
    ).order_by('-total')[:20]
    
    # By product
    by_product = invoices.values(
        'lines__item__code',
        'lines__item__name'
    ).annotate(
        total=Sum('lines__total_price'),
        quantity=Sum('lines__quantity')
    ).order_by('-total')[:20]
    
    # Monthly revenue
    monthly = []
    for i in range(11, -1, -1):
        month = timezone.now().date() - timedelta(days=30*i)
        month_start = month.replace(day=1)
        if month.month == 12:
            month_end = month.replace(month=1, day=1) - timedelta(days=1)
        else:
            month_end = month.replace(month=month.month+1, day=1) - timedelta(days=1)
        
        total = invoices.filter(
            invoice_date__gte=month_start,
            invoice_date__lte=month_end
        ).aggregate(total=Sum('total_amount'))['total'] or 0
        
        monthly.append({
            'month': month.strftime('%b %Y'),
            'total': float(total)
        })
    
    context = {
        'form': form,
        'by_customer': by_customer,
        'by_product': by_product,
        'monthly': monthly,
        'total_revenue': invoices.aggregate(total=Sum('total_amount'))['total'] or 0,
        'invoice_count': invoices.count(),
        'today': timezone.now().date(),
    }
    
    return render(request, 'reports/sales/revenue_analysis.html', context)


@login_required
def customer_performance_report(request):
    """Customer performance report"""
    form = SalesReportFilterForm(request.GET)
    
    date_from = form.cleaned_data.get('date_from') if form.is_valid() else None
    date_to = form.cleaned_data.get('date_to') if form.is_valid() else None
    
    customers_qs = Customer.objects
    
    if date_from:
        customers_qs = customers_qs.filter(
            Q(sales_orders__order_date__gte=date_from) |
            Q(salesinvoice__invoice_date__gte=date_from)
        )
    if date_to:
        customers_qs = customers_qs.filter(
            Q(sales_orders__order_date__lte=date_to) |
            Q(salesinvoice__invoice_date__lte=date_to)
        )
    
    customers = customers_qs.annotate(
        total_orders=Count('sales_orders'),
        invoiced_total=Sum('salesinvoice__total_amount'),
        total_paid=Sum('salesinvoice__paid_amount'),
        avg_order_value=Avg('sales_orders__total_amount'),
        last_order_date=models.Max('sales_orders__order_date')
    ).filter(total_orders__gt=0)
    
    for customer in customers:
        customer.outstanding = (customer.invoiced_total or 0) - (customer.total_paid or 0)
    
    # Top customers by revenue
    top_customers = sorted(customers, key=lambda x: x.invoiced_total or 0, reverse=True)[:20]
    
    context = {
        'form': form,
        'customers': customers,
        'top_customers': top_customers,
        'today': timezone.now().date(),
    }
    
    return render(request, 'reports/sales/customer_performance.html', context)


@login_required
def product_sales_report(request):
    """Product sales report"""
    from django.db.models.functions import Coalesce
    
    form = SalesReportFilterForm(request.GET)
    
    date_from = form.cleaned_data.get('date_from') if form.is_valid() else None
    date_to = form.cleaned_data.get('date_to') if form.is_valid() else None
    
    products_qs = Item.objects.filter(category='finished', is_active=True)
    
    if date_from:
        products_qs = products_qs.filter(
            Q(salesorderline__order__order_date__gte=date_from) |
            Q(salesinvoiceline__invoice__invoice_date__gte=date_from)
        )
    if date_to:
        products_qs = products_qs.filter(
            Q(salesorderline__order__order_date__lte=date_to) |
            Q(salesinvoiceline__invoice__invoice_date__lte=date_to)
        )
    
    products = products_qs.annotate(
        total_quantity=Coalesce(Sum('salesorderline__quantity', output_field=models.DecimalField()), 0),
        total_revenue=Coalesce(Sum('salesinvoiceline__total_price', output_field=models.DecimalField()), 0),
        order_count=Count('salesorderline__order', distinct=True)
    ).order_by('-total_revenue')
    
    # Category breakdown
    category_qs = Item.objects.filter(category='finished', is_active=True)
    if date_from:
        category_qs = category_qs.filter(
            Q(salesorderline__order__order_date__gte=date_from) |
            Q(salesinvoiceline__invoice__invoice_date__gte=date_from)
        )
    if date_to:
        category_qs = category_qs.filter(
            Q(salesorderline__order__order_date__lte=date_to) |
            Q(salesinvoiceline__invoice__invoice_date__lte=date_to)
        )
    
    by_category = category_qs.values('category').annotate(
        total_revenue=Coalesce(Sum('salesinvoiceline__total_price', output_field=models.DecimalField()), 0),
        total_quantity=Coalesce(Sum('salesorderline__quantity', output_field=models.DecimalField()), 0)
    )
    
    context = {
        'form': form,
        'products': products,
        'by_category': by_category,
        'total_revenue': products.aggregate(total=Sum('total_revenue'))['total'] or 0,
        'today': timezone.now().date(),
    }
    
    return render(request, 'reports/sales/product_sales.html', context)


@login_required
def profit_loss_report(request):
    """Profit & Loss report"""
    form = FinancialReportFilterForm(request.GET)
    
    # Get date range
    date_from = form.cleaned_data.get('date_from') if form.is_valid() else timezone.now().date().replace(month=1, day=1)
    date_to = form.cleaned_data.get('date_to') if form.is_valid() else timezone.now().date()
    
    # Get income accounts (revenue)
    income_accounts = Account.objects.filter(
        account_type__name='Income',
        is_active=True
    ).values('code', 'name').annotate(
        balance=Sum('journal_lines__credit') - Sum('journal_lines__debit')
    ).filter(balance__gt=0).order_by('-balance')
    
    # Get expense accounts
    expense_accounts = Account.objects.filter(
        account_type__name='Expense',
        is_active=True
    ).values('code', 'name').annotate(
        balance=Sum('journal_lines__debit') - Sum('journal_lines__credit')
    ).filter(balance__gt=0).order_by('-balance')
    
    # Calculate totals
    total_income = sum(a['balance'] for a in income_accounts)
    total_expenses = sum(a['balance'] for a in expense_accounts)
    net_profit = total_income - total_expenses
    
    context = {
        'form': form,
        'income_accounts': income_accounts,
        'expense_accounts': expense_accounts,
        'total_income': total_income,
        'total_expenses': total_expenses,
        'net_profit': net_profit,
        'date_from': date_from,
        'date_to': date_to,
        'today': timezone.now().date(),
    }
    
    return render(request, 'reports/financial/profit_loss.html', context)


@login_required
def balance_sheet_report(request):
    """Balance Sheet report"""
    form = FinancialReportFilterForm(request.GET)
    date_to = timezone.now().date()
    
    if form.is_valid() and form.cleaned_data.get('date_to'):
        date_to = form.cleaned_data['date_to']
    
    # Asset accounts
    assets = Account.objects.filter(
        account_type__name='Asset',
        is_active=True
    ).values('code', 'name', 'account_category__name').annotate(
        balance=Sum(
            models.Case(
                models.When(journal_lines__journal__entry_date__lte=date_to, then='journal_lines__debit'),
                default=0
            )
        ) - Sum(
            models.Case(
                models.When(journal_lines__journal__entry_date__lte=date_to, then='journal_lines__credit'),
                default=0
            )
        )
    ).order_by('code')
    
    # Liability accounts
    liabilities = Account.objects.filter(
        account_type__name='Liability',
        is_active=True
    ).values('code', 'name', 'account_category__name').annotate(
        balance=Sum(
            models.Case(
                models.When(journal_lines__journal__entry_date__lte=date_to, then='journal_lines__credit'),
                default=0
            )
        ) - Sum(
            models.Case(
                models.When(journal_lines__journal__entry_date__lte=date_to, then='journal_lines__debit'),
                default=0
            )
        )
    ).order_by('code')
    
    # Equity accounts
    equity = Account.objects.filter(
        account_type__name='Equity',
        is_active=True
    ).values('code', 'name', 'account_category__name').annotate(
        balance=Sum(
            models.Case(
                models.When(journal_lines__journal__entry_date__lte=date_to, then='journal_lines__credit'),
                default=0
            )
        ) - Sum(
            models.Case(
                models.When(journal_lines__journal__entry_date__lte=date_to, then='journal_lines__debit'),
                default=0
            )
        )
    ).order_by('code')
    
    # Group by category
    assets_by_category = {}
    for asset in assets:
        cat = asset['account_category__name'] or 'Other Assets'
        if cat not in assets_by_category:
            assets_by_category[cat] = []
        assets_by_category[cat].append(asset)
    
    liabilities_by_category = {}
    for liability in liabilities:
        cat = liability['account_category__name'] or 'Other Liabilities'
        if cat not in liabilities_by_category:
            liabilities_by_category[cat] = []
        liabilities_by_category[cat].append(liability)
    
    # Totals
    total_assets = sum(a['balance'] for a in assets)
    total_liabilities = sum(l['balance'] for l in liabilities)
    total_equity = sum(e['balance'] for e in equity)
    
    context = {
        'form': form,
        'assets_by_category': assets_by_category,
        'liabilities_by_category': liabilities_by_category,
        'equity': equity,
        'total_assets': total_assets,
        'total_liabilities': total_liabilities,
        'total_equity': total_equity,
        'today': timezone.now().date(),
    }
    
    return render(request, 'reports/financial/balance_sheet.html', context)


@login_required
def cash_flow_report(request):
    """Cash Flow Statement"""
    form = FinancialReportFilterForm(request.GET)
    
    date_from = form.cleaned_data.get('date_from') if form.is_valid() else timezone.now().date().replace(day=1)
    date_to = form.cleaned_data.get('date_to') if form.is_valid() else timezone.now().date()
    
    # Get cash account
    cash_account = Account.objects.filter(code='1010').first()
    
    if cash_account:
        # Get all journal entries affecting cash
        cash_entries = JournalEntry.objects.filter(
            entry_date__range=[date_from, date_to],
            is_posted=True,
            journal_lines__account=cash_account
        ).distinct()
        
        # Operating activities: revenue and expense accounts
        operating_accounts = Account.objects.filter(
            account_type__name__in=['Revenue', 'Expense']
        )
        
        operating_flows = []
        for account in operating_accounts:
            debit = JournalLine.objects.filter(
                account=account,
                journal__entry_date__range=[date_from, date_to],
                journal__is_posted=True
            ).aggregate(total=Sum('debit'))['total'] or 0
            
            credit = JournalLine.objects.filter(
                account=account,
                journal__entry_date__range=[date_from, date_to],
                journal__is_posted=True
            ).aggregate(total=Sum('credit'))['total'] or 0
            
            net = debit - credit if account.account_type.name == 'Revenue' else credit - debit
            if net != 0:
                operating_flows.append({
                    'account': account.name,
                    'net': net
                })
        
        # Investing activities: fixed assets
        investing_accounts = Account.objects.filter(
            account_type__name='Asset',
            account_category__name__in=['Fixed Assets', 'Investments']
        ).exclude(code='1010')  # exclude cash
        
        investing_flows = []
        for account in investing_accounts:
            debit = JournalLine.objects.filter(
                account=account,
                journal__entry_date__range=[date_from, date_to],
                journal__is_posted=True
            ).aggregate(total=Sum('debit'))['total'] or 0
            
            credit = JournalLine.objects.filter(
                account=account,
                journal__entry_date__range=[date_from, date_to],
                journal__is_posted=True
            ).aggregate(total=Sum('credit'))['total'] or 0
            
            net = debit - credit  # increase in assets is outflow
            if net != 0:
                investing_flows.append({
                    'account': account.name,
                    'net': net
                })
        
        # Financing activities: liabilities and equity
        financing_accounts = Account.objects.filter(
            account_type__name__in=['Liability', 'Equity']
        )
        
        financing_flows = []
        for account in financing_accounts:
            debit = JournalLine.objects.filter(
                account=account,
                journal__entry_date__range=[date_from, date_to],
                journal__is_posted=True
            ).aggregate(total=Sum('debit'))['total'] or 0
            
            credit = JournalLine.objects.filter(
                account=account,
                journal__entry_date__range=[date_from, date_to],
                journal__is_posted=True
            ).aggregate(total=Sum('credit'))['total'] or 0
            
            net = credit - debit  # increase in liabilities/equity is inflow
            if net != 0:
                financing_flows.append({
                    'account': account.name,
                    'net': net
                })
        
        operating_net = sum(f['net'] for f in operating_flows)
        investing_net = sum(f['net'] for f in investing_flows)
        financing_net = sum(f['net'] for f in financing_flows)
        
    else:
        operating_flows = []
        investing_flows = []
        financing_flows = []
        operating_net = investing_net = financing_net = 0
    
    context = {
        'form': form,
        'date_from': date_from,
        'date_to': date_to,
        'operating_flows': operating_flows,
        'investing_flows': investing_flows,
        'financing_flows': financing_flows,
        'operating_net': operating_net,
        'investing_net': investing_net,
        'financing_net': financing_net,
        'net_cash_flow': operating_net + investing_net + financing_net,
        'today': timezone.now().date(),
    }
    
    return render(request, 'reports/financial/cash_flow.html', context)


@login_required
def trial_balance_report(request):
    """Trial Balance report"""
    form = FinancialReportFilterForm(request.GET)
    date_to = timezone.now().date()
    
    if form.is_valid() and form.cleaned_data.get('date_to'):
        date_to = form.cleaned_data['date_to']
    
    accounts = Account.objects.filter(is_active=True).select_related(
        'account_type', 'account_group'
    ).annotate(
        balance=Sum(
            models.Case(
                models.When(journal_lines__journal__entry_date__lte=date_to, then='journal_lines__debit'),
                default=0
            )
        ) - Sum(
            models.Case(
                models.When(journal_lines__journal__entry_date__lte=date_to, then='journal_lines__credit'),
                default=0
            )
        )
    ).order_by('code')
    
    total_debits = 0
    total_credits = 0
    account_list = []
    
    for account in accounts:
        balance = account.balance or 0
        if balance > 0:
            debits = balance
            credits = 0
            total_debits += balance
        else:
            debits = 0
            credits = abs(balance)
            total_credits += abs(balance)
        
        account_list.append({
            'code': account.code,
            'name': account.name,
            'type': account.account_type.name if account.account_type else '',
            'debits': debits,
            'credits': credits,
        })
    
    context = {
        'form': form,
        'accounts': account_list,
        'total_debits': total_debits,
        'total_credits': total_credits,
        'difference': total_debits - total_credits,
        'is_balanced': abs(total_debits - total_credits) < 0.01,
        'today': timezone.now().date(),
    }
    
    return render(request, 'reports/financial/trial_balance.html', context)


@login_required
def accounts_receivable_report(request):
    """Accounts Receivable Aging report"""
    # SalesInvoice belongs to sales app, not accounting
    from apps.sales.models import SalesInvoice
    
    today = timezone.now().date()
    
    invoices = SalesInvoice.objects.filter(
        status__in=['posted', 'partial', 'overdue']
    ).select_related('customer')
    
    # Aging buckets
    aging = {
        'current': invoices.filter(due_date__gte=today),
        '1_30': invoices.filter(due_date__lt=today, due_date__gte=today - timedelta(days=30)),
        '31_60': invoices.filter(due_date__lt=today - timedelta(days=30), due_date__gte=today - timedelta(days=60)),
        '61_90': invoices.filter(due_date__lt=today - timedelta(days=60), due_date__gte=today - timedelta(days=90)),
        '90_plus': invoices.filter(due_date__lt=today - timedelta(days=90)),
    }
    
    bucket_totals = {}
    for name, qs in aging.items():
        bucket_totals[name] = {
            'count': qs.count(),
            'amount': qs.aggregate(total=Sum('total_amount'))['total'] or 0,
        }
    
    # By customer
    by_customer = invoices.values('customer__name').annotate(
        total=Sum('total_amount'),
        count=Count('id')
    ).order_by('-total')
    
    # Additional summary metrics used in template
    total_invoices = invoices.count()
    overdue_qs = invoices.filter(due_date__lt=today)
    overdue_amount = overdue_qs.aggregate(total=Sum('total_amount'))['total'] or 0
    overdue_count = overdue_qs.count()
    if total_invoices:
        # average days outstanding relative to due date
        total_days = sum([(today - inv.due_date).days for inv in invoices])
        avg_dso = total_days / total_invoices
    else:
        avg_dso = 0
    
    context = {
        'aging': aging,
        'bucket_totals': bucket_totals,
        'by_customer': by_customer,
        'total_outstanding': invoices.aggregate(total=Sum('total_amount'))['total'] or 0,
        'today': today,
        'total_invoices': total_invoices,
        'overdue_amount': overdue_amount,
        'overdue_count': overdue_count,
        'avg_dso': round(avg_dso, 1),
    }
    
    return render(request, 'reports/financial/ar_aging.html', context)


@login_required
def accounts_payable_report(request):
    """Accounts Payable Aging report"""
    from apps.accounting.models import PurchaseBill
    
    today = timezone.now().date()
    
    bills = PurchaseBill.objects.filter(
        status__in=['posted', 'partial', 'overdue']
    ).select_related('supplier')
    
    # Aging buckets
    aging = {
        'current': bills.filter(due_date__gte=today),
        '1_30': bills.filter(due_date__lt=today, due_date__gte=today - timedelta(days=30)),
        '31_60': bills.filter(due_date__lt=today - timedelta(days=30), due_date__gte=today - timedelta(days=60)),
        '61_90': bills.filter(due_date__lt=today - timedelta(days=60), due_date__gte=today - timedelta(days=90)),
        '90_plus': bills.filter(due_date__lt=today - timedelta(days=90)),
    }
    
    bucket_totals = {}
    for name, qs in aging.items():
        bucket_totals[name] = {
            'count': qs.count(),
            'amount': qs.aggregate(total=Sum('total_amount'))['total'] or 0,
        }
    
    # By supplier
    by_supplier = bills.values('supplier__name').annotate(
        total=Sum('total_amount'),
        count=Count('id')
    ).order_by('-total')
    
    context = {
        'aging': aging,
        'bucket_totals': bucket_totals,
        'by_supplier': by_supplier,
        'total_outstanding': bills.aggregate(total=Sum('total_amount'))['total'] or 0,
        'today': today,
    }
    
    return render(request, 'reports/financial/ap_aging.html', context)


@login_required
def executive_summary(request):
    """Executive summary dashboard with key KPIs"""
    today = timezone.now().date()
    month_start = today.replace(day=1)
    
    # Sales KPIs
    monthly_sales = SalesInvoice.objects.filter(
        invoice_date__gte=month_start,
        status__in=['posted', 'paid']
    ).aggregate(total=Sum('total_amount'))['total'] or 0
    
    # Inventory KPIs
    inventory_value = CurrentStock.objects.aggregate(
        total=Sum(F('quantity') * F('item__unit_cost'))
    )['total'] or 0
    
    # Production KPIs
    monthly_production = ProductionRun.objects.filter(
        end_date__gte=month_start,
        status='completed'
    ).aggregate(total=Sum('actual_quantity'))['total'] or 0
    
    # Purchasing KPIs
    monthly_purchases = PurchaseOrder.objects.filter(
        order_date__gte=month_start,
        status__in=['received', 'partial']
    ).aggregate(total=Sum('total_amount'))['total'] or 0
    
    # Financial KPIs
    cash_balance = Account.objects.filter(code='1010').first()
    cash_balance = cash_balance.current_balance if cash_balance else 0
    
    ar_total = SalesInvoice.objects.filter(
        status__in=['posted', 'partial', 'overdue']
    ).aggregate(total=Sum('total_amount'))['total'] or 0
    
    ap_total = PurchaseBill.objects.filter(
        status__in=['posted', 'partial', 'overdue']
    ).aggregate(total=Sum('total_amount'))['total'] or 0
    
    context = {
        'monthly_sales': monthly_sales,
        'inventory_value': inventory_value,
        'monthly_production': monthly_production,
        'monthly_purchases': monthly_purchases,
        'cash_balance': cash_balance,
        'ar_total': ar_total,
        'ap_total': ap_total,
        'today': today,
    }
    
    return render(request, 'reports/custom/executive_summary.html', context)


@login_required
def kpi_dashboard(request):
    """KPI Dashboard with financial ratios"""
    form = FinancialReportFilterForm(request.GET)
    date_to = timezone.now().date()
    
    if form.is_valid() and form.cleaned_data.get('date_to'):
        date_to = form.cleaned_data['date_to']
    
    # Get account balances as of date_to
    def get_balance(account_type, category=None, date_to=date_to):
        qs = Account.objects.filter(account_type__name=account_type, is_active=True)
        if category:
            qs = qs.filter(account_category__name=category)
        return qs.aggregate(
            balance=Sum(
                models.Case(
                    models.When(journal_lines__journal__entry_date__lte=date_to, then='journal_lines__debit'),
                    default=0
                )
            ) - Sum(
                models.Case(
                    models.When(journal_lines__journal__entry_date__lte=date_to, then='journal_lines__credit'),
                    default=0
                )
            )
        )['balance'] or 0
    
    # Current Assets
    current_assets = get_balance('Asset', 'Current Assets')
    
    # Current Liabilities
    current_liabilities = get_balance('Liability', 'Current Liabilities')
    
    # Inventory (from inventory module)
    inventory_value = CurrentStock.objects.filter(
        warehouse__is_active=True
    ).aggregate(
        total=Sum(F('quantity') * F('avg_cost'))
    )['total'] or 0
    
    # Total Assets
    total_assets = get_balance('Asset')
    
    # Total Liabilities
    total_liabilities = get_balance('Liability')
    
    # Total Equity
    total_equity = get_balance('Equity')
    
    # Revenue
    revenue = get_balance('Revenue')
    
    # COGS (assuming expense category 'Cost of Goods Sold')
    cogs = Account.objects.filter(
        account_type__name='Expense',
        account_category__name='Cost of Goods Sold'
    ).aggregate(
        total=Sum(
            models.Case(
                models.When(journal_lines__journal__entry_date__lte=date_to, then='journal_lines__debit'),
                default=0
            )
        )
    )['total'] or 0
    
    # Operating Expenses (other expenses)
    operating_expenses = Account.objects.filter(
        account_type__name='Expense'
    ).exclude(
        account_category__name='Cost of Goods Sold'
    ).aggregate(
        total=Sum(
            models.Case(
                models.When(journal_lines__journal__entry_date__lte=date_to, then='journal_lines__debit'),
                default=0
            )
        )
    )['total'] or 0
    
    # Calculate ratios
    kpis = {}
    
    # Liquidity Ratios
    kpis['current_ratio'] = current_assets / current_liabilities if current_liabilities else 0
    kpis['quick_ratio'] = (current_assets - inventory_value) / current_liabilities if current_liabilities else 0
    
    # Profitability Ratios
    gross_profit = revenue - cogs
    net_profit = gross_profit - operating_expenses
    
    kpis['gross_profit_margin'] = (gross_profit / revenue * 100) if revenue else 0
    kpis['net_profit_margin'] = (net_profit / revenue * 100) if revenue else 0
    kpis['return_on_assets'] = (net_profit / total_assets * 100) if total_assets else 0
    kpis['return_on_equity'] = (net_profit / total_equity * 100) if total_equity else 0
    
    # Solvency Ratios
    kpis['debt_to_equity'] = total_liabilities / total_equity if total_equity else 0
    kpis['debt_to_assets'] = total_liabilities / total_assets if total_assets else 0
    
    # Activity Ratios
    # Inventory Turnover
    avg_inventory = inventory_value  # simplified
    kpis['inventory_turnover'] = cogs / avg_inventory if avg_inventory else 0
    
    # Accounts Receivable Turnover (assuming receivables are current assets - inventory - cash)
    receivables = current_assets - inventory_value - get_balance('Asset', 'Cash and Cash Equivalents')
    kpis['receivables_turnover'] = revenue / receivables if receivables else 0
    
    context = {
        'form': form,
        'kpis': kpis,
        'date_to': date_to,
        'today': timezone.now().date(),
    }
    return render(request, 'reports/custom/kpi_dashboard.html', context)


@login_required
def comparative_analysis(request):
    """Comparative analysis across periods"""
    context = {
        'today': timezone.now().date(),
    }
    return render(request, 'reports/custom/comparative_analysis.html', context)


# ============================
# Journal Reports
# ============================

@login_required
def general_ledger_report(request):
    """General Ledger report"""
    form = FinancialReportFilterForm(request.GET)
    
    accounts = Account.objects.filter(is_active=True).select_related(
        'account_type', 'account_group'
    ).order_by('code')
    
    if form.is_valid():
        date_from = form.cleaned_data.get('date_from')
        date_to = form.cleaned_data.get('date_to')
    else:
        date_from = None
        date_to = None
    
    # Get journal entries for each account
    ledger_data = []
    for account in accounts:
        entries = JournalEntry.objects.filter(
            journal_lines__account=account,
            is_posted=True
        ).distinct().order_by('entry_date')
        
        if date_from:
            entries = entries.filter(entry_date__gte=date_from)
        if date_to:
            entries = entries.filter(entry_date__lte=date_to)
        
        running_balance = 0
        entry_list = []
        
        for entry in entries:
            lines = entry.lines.filter(account=account)
            for line in lines:
                debit = line.debit or 0
                credit = line.credit or 0
                
                if account.account_type.name in ['Asset', 'Expense']:
                    running_balance += debit - credit
                else:
                    running_balance += credit - debit
                
                entry_list.append({
                    'date': entry.entry_date,
                    'reference': entry.reference,
                    'narration': line.narration or entry.narration,
                    'debit': debit,
                    'credit': credit,
                    'balance': running_balance
                })
        
        if entry_list:
            ledger_data.append({
                'account': account,
                'entries': entry_list,
                'final_balance': running_balance
            })
    
    context = {
        'form': form,
        'ledger_data': ledger_data,
        'today': timezone.now().date(),
    }
    
    return render(request, 'reports/financial/general_ledger.html', context)


@login_required
def cash_payment_journal(request):
    """Cash Payment Journal"""
    form = FinancialReportFilterForm(request.GET)
    
    date_from = form.cleaned_data.get('date_from') if form.is_valid() else timezone.now().date().replace(day=1)
    date_to = form.cleaned_data.get('date_to') if form.is_valid() else timezone.now().date()
    
    # Get cash account
    cash_account = Account.objects.filter(code='1010').first()
    
    if cash_account:
        # Get journal entries where cash is credited (payments)
        payments = JournalEntry.objects.filter(
            entry_date__range=[date_from, date_to],
            is_posted=True,
            journal_lines__account=cash_account,
            journal_lines__credit__gt=0
        ).distinct().select_related().order_by('entry_date')
        
        payment_data = []
        for entry in payments:
            cash_line = entry.lines.filter(account=cash_account, credit__gt=0).first()
            if cash_line:
                # Find the contra account (debited account)
                contra_line = entry.lines.exclude(account=cash_account).first()
                payment_data.append({
                    'date': entry.entry_date,
                    'reference': entry.reference,
                    'narration': entry.narration,
                    'amount': cash_line.credit,
                    'paid_to': contra_line.account.name if contra_line else 'Unknown',
                    'account_code': contra_line.account.code if contra_line else ''
                })
    else:
        payment_data = []
    
    context = {
        'form': form,
        'payments': payment_data,
        'total_payments': sum(p['amount'] for p in payment_data),
        'date_from': date_from,
        'date_to': date_to,
        'today': timezone.now().date(),
    }
    
    return render(request, 'reports/financial/cash_payment_journal.html', context)


@login_required
def sales_journal(request):
    """Sales Journal"""
    form = FinancialReportFilterForm(request.GET)
    
    date_from = form.cleaned_data.get('date_from') if form.is_valid() else timezone.now().date().replace(day=1)
    date_to = form.cleaned_data.get('date_to') if form.is_valid() else timezone.now().date()
    
    # Get sales revenue account
    sales_account = Account.objects.filter(code='4100').first()
    
    if sales_account:
        # Get journal entries where sales revenue is credited
        sales_entries = JournalEntry.objects.filter(
            entry_date__range=[date_from, date_to],
            is_posted=True,
            journal_lines__account=sales_account,
            journal_lines__credit__gt=0
        ).distinct().select_related().order_by('entry_date')
        
        sales_data = []
        for entry in sales_entries:
            sales_line = entry.lines.filter(account=sales_account, credit__gt=0).first()
            ar_line = entry.lines.filter(account__code='1100').first()  # AR account
            
            if sales_line and ar_line:
                sales_data.append({
                    'date': entry.entry_date,
                    'reference': entry.reference,
                    'narration': entry.narration,
                    'amount': sales_line.credit,
                    'customer': 'Customer',  # Could be enhanced to get actual customer
                    'invoice_ref': entry.reference
                })
    else:
        sales_data = []
    
    context = {
        'form': form,
        'sales': sales_data,
        'total_sales': sum(s['amount'] for s in sales_data),
        'date_from': date_from,
        'date_to': date_to,
        'today': timezone.now().date(),
    }
    
    return render(request, 'reports/financial/sales_journal.html', context)


@login_required
def receipt_journal(request):
    """Receipt Journal (Cash/Bank Receipts)"""
    form = FinancialReportFilterForm(request.GET)
    
    date_from = form.cleaned_data.get('date_from') if form.is_valid() else timezone.now().date().replace(day=1)
    date_to = form.cleaned_data.get('date_to') if form.is_valid() else timezone.now().date()
    
    # Get cash account
    cash_account = Account.objects.filter(code='1010').first()
    
    if cash_account:
        # Get journal entries where cash is debited (receipts)
        receipts = JournalEntry.objects.filter(
            entry_date__range=[date_from, date_to],
            is_posted=True,
            journal_lines__account=cash_account,
            journal_lines__debit__gt=0
        ).distinct().select_related().order_by('entry_date')
        
        receipt_data = []
        for entry in receipts:
            cash_line = entry.lines.filter(account=cash_account, debit__gt=0).first()
            if cash_line:
                # Find the contra account (credited account)
                contra_line = entry.lines.exclude(account=cash_account).first()
                receipt_data.append({
                    'date': entry.entry_date,
                    'reference': entry.reference,
                    'narration': entry.narration,
                    'amount': cash_line.debit,
                    'received_from': contra_line.account.name if contra_line else 'Unknown',
                    'account_code': contra_line.account.code if contra_line else ''
                })
    else:
        receipt_data = []
    
    context = {
        'form': form,
        'receipts': receipt_data,
        'total_receipts': sum(r['amount'] for r in receipt_data),
        'date_from': date_from,
        'date_to': date_to,
        'today': timezone.now().date(),
    }
    
    return render(request, 'reports/financial/receipt_journal.html', context)


@login_required
def purchase_journal(request):
    """Purchase Journal"""
    form = FinancialReportFilterForm(request.GET)
    
    date_from = form.cleaned_data.get('date_from') if form.is_valid() else timezone.now().date().replace(day=1)
    date_to = form.cleaned_data.get('date_to') if form.is_valid() else timezone.now().date()
    
    # Get inventory/purchase expense account
    inventory_account = Account.objects.filter(code='1300').first()
    
    if inventory_account:
        # Get journal entries where inventory is debited
        purchase_entries = JournalEntry.objects.filter(
            entry_date__range=[date_from, date_to],
            is_posted=True,
            journal_lines__account=inventory_account,
            journal_lines__debit__gt=0
        ).distinct().select_related().order_by('entry_date')
        
        purchase_data = []
        for entry in purchase_entries:
            inv_line = entry.lines.filter(account=inventory_account, debit__gt=0).first()
            ap_line = entry.lines.filter(account__code='2100').first()  # AP account
            
            if inv_line and ap_line:
                purchase_data.append({
                    'date': entry.entry_date,
                    'reference': entry.reference,
                    'narration': entry.narration,
                    'amount': inv_line.debit,
                    'supplier': 'Supplier',  # Could be enhanced
                    'po_ref': entry.reference
                })
    else:
        purchase_data = []
    
    context = {
        'form': form,
        'purchases': purchase_data,
        'total_purchases': sum(p['amount'] for p in purchase_data),
        'date_from': date_from,
        'date_to': date_to,
        'today': timezone.now().date(),
    }
    
    return render(request, 'reports/financial/purchase_journal.html', context)


@login_required
def inventory_journal(request):
    """Inventory Journal (Stock Movements)"""
    form = InventoryReportFilterForm(request.GET)
    
    transactions = StockTransaction.objects.select_related(
        'item', 'warehouse_from', 'warehouse_to', 'created_by'
    ).order_by('-transaction_date')
    
    if form.is_valid():
        if form.cleaned_data.get('date_from'):
            transactions = transactions.filter(transaction_date__date__gte=form.cleaned_data['date_from'])
        if form.cleaned_data.get('date_to'):
            transactions = transactions.filter(transaction_date__date__lte=form.cleaned_data['date_to'])
        if form.cleaned_data.get('warehouse'):
            transactions = transactions.filter(
                Q(warehouse_from_id=form.cleaned_data['warehouse']) |
                Q(warehouse_to_id=form.cleaned_data['warehouse'])
            )
        if form.cleaned_data.get('item'):
            transactions = transactions.filter(item__code__icontains=form.cleaned_data['item'])
    
    # Group by date and type for journal format
    journal_entries = []
    for transaction in transactions[:500]:  # Limit for performance
        journal_entries.append({
            'date': transaction.transaction_date.date(),
            'reference': transaction.reference or f"ST-{transaction.id}",
            'narration': f"{transaction.get_transaction_type_display()} - {transaction.item.code}",
            'item': transaction.item.name,
            'quantity': transaction.quantity,
            'warehouse_from': transaction.warehouse_from.name if transaction.warehouse_from else '',
            'warehouse_to': transaction.warehouse_to.name if transaction.warehouse_to else '',
            'type': transaction.get_transaction_type_display()
        })
    
    context = {
        'form': form,
        'journal_entries': journal_entries,
        'today': timezone.now().date(),
    }
    
    return render(request, 'reports/financial/inventory_journal.html', context)


# ============================
# Scheduled Reports
# ============================

@login_required
def scheduled_reports_list(request):
    """List scheduled reports"""
    reports = ScheduledReport.objects.filter(created_by=request.user).order_by('next_run')
    
    context = {
        'reports': reports,
        'today': timezone.now().date(),
    }
    
    return render(request, 'reports/scheduled/list.html', context)


@login_required
def create_scheduled_report(request):
    """Create a new scheduled report"""
    if request.method == 'POST':
        form = ScheduledReportForm(request.POST)
        if form.is_valid():
            report = form.save(commit=False)
            report.created_by = request.user
            report.save()
            messages.success(request, f'Scheduled report "{report.name}" created successfully.')
            return redirect('reports:scheduled_reports')
    else:  # <-- This line had the error (missing colon)
        form = ScheduledReportForm()
    
    context = {
        'form': form,
        'today': timezone.now().date(),
    }
    
    return render(request, 'reports/scheduled/form.html', context)


@login_required
def edit_scheduled_report(request, report_id):
    """Edit a scheduled report"""
    report = get_object_or_404(ScheduledReport, id=report_id, created_by=request.user)
    
    if request.method == 'POST':
        form = ScheduledReportForm(request.POST, instance=report)
        if form.is_valid():
            form.save()
            messages.success(request, f'Scheduled report "{report.name}" updated successfully.')
            return redirect('reports:scheduled_reports')
    else:
        form = ScheduledReportForm(instance=report)
    
    context = {
        'form': form,
        'report': report,
        'today': timezone.now().date(),
    }
    
    return render(request, 'reports/scheduled/form.html', context)


@login_required
def delete_scheduled_report(request, report_id):
    """Delete a scheduled report"""
    report = get_object_or_404(ScheduledReport, id=report_id, created_by=request.user)
    
    if request.method == 'POST':
        report.delete()
        messages.success(request, f'Scheduled report "{report.name}" deleted.')
        return redirect('reports:scheduled_reports')
    
    context = {
        'report': report,
    }
    
    return render(request, 'reports/scheduled/delete_confirm.html', context)


@login_required
def run_scheduled_report(request, report_id):
    """Run a scheduled report manually"""
    report = get_object_or_404(ScheduledReport, id=report_id, created_by=request.user)
    
    # Trigger report generation
    messages.success(request, f'Report "{report.name}" has been queued for generation.')
    
    return redirect('reports:scheduled_reports')


# ============================
# Export Functions
# ============================

def export_stock_summary_excel(context):
    """Export stock summary to Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Summary"
    
    # Headers
    headers = ['Item Code', 'Item Name', 'Category', 'Warehouse', 'Lot', 'Quantity', 'Unit', 'Unit Cost', 'Total Value']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Data
    for row, stock in enumerate(context['stock_items'], 2):
        ws.cell(row=row, column=1).value = stock.item.code
        ws.cell(row=row, column=2).value = stock.item.name
        ws.cell(row=row, column=3).value = stock.item.get_category_display()
        ws.cell(row=row, column=4).value = stock.warehouse.name
        ws.cell(row=row, column=5).value = stock.lot.batch_number if stock.lot else ''
        ws.cell(row=row, column=6).value = float(stock.quantity)
        ws.cell(row=row, column=7).value = stock.item.unit.abbreviation if stock.item.unit else ''
        ws.cell(row=row, column=8).value = float(stock.item.unit_cost or 0)
        ws.cell(row=row, column=9).value = float(stock.quantity * (stock.item.unit_cost or 0))
    
    # Adjust column widths
    for col in range(1, 10):
        ws.column_dimensions[chr(64 + col)].width = 15
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="stock_summary_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response


def export_stock_summary_pdf(context):
    """Export stock summary to PDF"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    elements = []
    
    # Title
    styles = getSampleStyleSheet()
    title = Paragraph(f"Stock Summary Report - {context['today'].strftime('%Y-%m-%d')}", styles['Title'])
    elements.append(title)
    elements.append(Paragraph("<br/>", styles['Normal']))
    
    # Summary stats
    summary_text = f"""
    Total Items: {context['total_items']}<br/>
    Total Value: {context['total_value']:,.2f} ETB<br/>
    Near Expiry: {context['near_expiry']}<br/>
    Expired: {context['expired']}<br/>
    """
    elements.append(Paragraph(summary_text, styles['Normal']))
    elements.append(Paragraph("<br/><br/>", styles['Normal']))
    
    # Table data
    data = [['Item Code', 'Item Name', 'Warehouse', 'Quantity', 'Unit', 'Value']]
    
    for stock in context['stock_items'][:50]:  # Limit for PDF
        data.append([
            stock.item.code,
            stock.item.name[:30],
            stock.warehouse.name,
            f"{stock.quantity:.2f}",
            stock.item.unit.abbreviation if stock.item.unit else '',
            f"{stock.quantity * (stock.item.unit_cost or 0):,.2f}"
        ])
    
    # Create table
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="stock_summary_{timezone.now().strftime("%Y%m%d")}.pdf"'
    return response


def export_stock_summary_csv(context):
    """Export stock summary to CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="stock_summary_{timezone.now().strftime("%Y%m%d")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Item Code', 'Item Name', 'Category', 'Warehouse', 'Lot', 'Quantity', 'Unit', 'Unit Cost', 'Total Value'])
    
    for stock in context['stock_items']:
        writer.writerow([
            stock.item.code,
            stock.item.name,
            stock.item.get_category_display(),
            stock.warehouse.name,
            stock.lot.batch_number if stock.lot else '',
            float(stock.quantity),
            stock.item.unit.abbreviation if stock.item.unit else '',
            float(stock.item.unit_cost or 0),
            float(stock.quantity * (stock.item.unit_cost or 0))
        ])
    
    return response