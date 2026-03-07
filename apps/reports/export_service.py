import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse

class ExcelExportService:
    @staticmethod
    def export_to_excel(filename, columns, data):
        # Create a workbook and select active sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report Data"

        # Style for the header
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        # Write Headers
        for col_num, column_title in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = column_title
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Write Data Rows
        for row_num, row_data in enumerate(data, 2):
            # If row_data is a dict (from .values()), we map by keys
            # If it's a list/tuple, we enumerate directly
            if isinstance(row_data, dict):
                values = list(row_data.values())
            else:
                values = row_data

            for col_num, cell_value in enumerate(values, 1):
                ws.cell(row=row_num, column=col_num).value = str(cell_value)

        # Auto-adjust column width
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 2

        # Prepare Response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename={filename}.xlsx'
        wb.save(response)
        return response